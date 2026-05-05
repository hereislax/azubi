# SPDX-License-Identifier: EUPL-1.2
# SPDX-FileCopyrightText: 2026 devNicoLax
"""Views für die Inventarverwaltung (Kategorien, Gegenstände, Ausgaben, Quittungen, Scan-Upload)."""
from datetime import date
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.http import require_POST

from services.paperless import PaperlessService
from student.models import Student

from .forms import InventoryImportForm
from .models import InventoryCategory, InventoryIssuance, InventoryItem, ReceiptTemplate

QR_PREFIX = "AZUBI-INV-"


def _make_qr_image(data: str) -> BytesIO:
    """Erzeugt ein QR-Code-Bild als PNG im BytesIO-Puffer."""
    import qrcode
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=6, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def _read_qr_from_pdf(pdf_bytes: bytes) -> str | None:
    """Rendert jede PDF-Seite und liest den ersten passenden QR-Code."""
    import fitz  # PyMuPDF
    import zxingcpp
    from PIL import Image

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for page in doc:
        mat = fitz.Matrix(3, 3)  # 3× Zoom für bessere Lesbarkeit
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        for result in zxingcpp.read_barcodes(img):
            if result.text.startswith(QR_PREFIX):
                return result.text
    return None


def _can_manage(user):
    """Prüft ob der Nutzer Inventar verwalten darf (Leitung oder Referat)."""
    from services.roles import is_training_director, is_training_office
    return is_training_director(user) or is_training_office(user)


def _require_manage(request):
    """Wirft PermissionDenied wenn der Nutzer kein Inventar-Verwaltungsrecht hat."""
    if not _can_manage(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied


# ── Übersicht ─────────────────────────────────────────────────────────────────

@login_required
def inventory_list(request):
    """Übersicht aller Inventargegenstände mit Filter nach Status, Kategorie und Suchbegriff."""
    _require_manage(request)
    categories = InventoryCategory.objects.prefetch_related("items").all()
    status_filter = request.GET.get("status", "")
    category_filter = request.GET.get("category", "")
    query = request.GET.get("q", "").strip()

    items = InventoryItem.objects.select_related("category").order_by("category__name", "name")
    if status_filter:
        items = items.filter(status=status_filter)
    if category_filter:
        items = items.filter(category_id=category_filter)
    if query:
        from django.db.models import Q
        items = items.filter(
            Q(name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(location__icontains=query)
        )

    return render(request, "inventory/inventory_list.html", {
        "items": items,
        "categories": categories,
        "status_choices": InventoryItem.Status.choices,
        "status_filter": status_filter,
        "category_filter": category_filter,
        "query": query,
    })


# ── QR-Code & Etiketten ───────────────────────────────────────────────────────

ITEM_QR_PREFIX = "AZUBI-INV-ITEM-"


def _item_qr_payload(item: "InventoryItem") -> str:
    """Inhalt des QR-Codes: Seriennummer, oder Fallback auf Item-Identifier."""
    return item.serial_number.strip() if item.serial_number else f"{ITEM_QR_PREFIX}{item.pk}"


@login_required
def item_qr_image(request, public_id):
    """Liefert ein QR-Code-PNG für einen einzelnen Gegenstand."""
    _require_manage(request)
    item = get_object_or_404(InventoryItem, public_id=public_id)
    buf = _make_qr_image(_item_qr_payload(item))
    return HttpResponse(buf.read(), content_type="image/png")


@login_required
def item_label(request, public_id):
    """Druckansicht: Einzelnes Etikett mit QR-Code, Seriennummer, Bezeichnung und Kategorie."""
    _require_manage(request)
    item = get_object_or_404(InventoryItem.objects.select_related("category"), public_id=public_id)
    return render(request, "inventory/labels_print.html", {
        "items": [item],
        "single": True,
    })


@login_required
def labels_print(request):
    """Druckansicht: Mehrere Etiketten als Bogen (Auswahl per ?ids=1,2,3 oder POST-Liste)."""
    _require_manage(request)
    raw_ids = request.GET.get("ids") or request.POST.get("ids", "")
    pks = [int(s) for s in raw_ids.split(",") if s.strip().isdigit()]
    if not pks:
        messages.error(request, "Bitte mindestens einen Gegenstand auswählen.")
        return redirect("inventory:inventory_list")
    items = list(
        InventoryItem.objects
        .select_related("category")
        .filter(pk__in=pks)
        .order_by("category__name", "name")
    )
    return render(request, "inventory/labels_print.html", {
        "items": items,
        "single": False,
    })


# ── Excel-Import ─────────────────────────────────────────────────────────────


@login_required
def inventory_import_template(request):
    """Erzeugt eine Excel-Vorlage mit Kopfzeile, Beispielzeile, Drop-downs und Referenzlisten."""
    _require_manage(request)

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import"
    headers = ["Bezeichnung", "Kategorie", "Seriennummer", "Status", "Lagerort", "Notizen"]
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([28, 22, 22, 16, 22, 36], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    example = ["Laptop ThinkPad T14", "<Kategorie aus Referenzliste>",
               "SN-12345", "Verfügbar", "Lager A", "Beispielzeile – kann gelöscht werden"]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(italic=True, color="888888")

    # Drop-down Status (Spalte D)
    status_labels = ",".join(label for _, label in InventoryItem.Status.choices)
    sdv = DataValidation(type="list", formula1=f'"{status_labels}"', allow_blank=True)
    sdv.error = "Bitte einen der vorgegebenen Werte wählen."
    sdv.errorTitle = "Ungültiger Status"
    ws.add_data_validation(sdv)
    sdv.add("D2:D1000")

    # Drop-down Kategorie (Spalte B) – nur wenn Werteliste ≤ 250 Zeichen passt
    categories = list(InventoryCategory.objects.order_by("name"))
    if categories:
        cat_values = ",".join(c.name for c in categories)
        if len(cat_values) <= 250:
            cdv = DataValidation(type="list", formula1=f'"{cat_values}"', allow_blank=False)
            ws.add_data_validation(cdv)
            cdv.add("B2:B1000")

    # ── Sheet „Kategorien (Referenz)" ─────────────────────────────────────
    ws_cat = wb.create_sheet("Kategorien (Referenz)")
    cat_headers = ["Bezeichnung", "Beschreibung"]
    for col_idx, h in enumerate(cat_headers, start=1):
        cell = ws_cat.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws_cat.column_dimensions["A"].width = 30
    ws_cat.column_dimensions["B"].width = 60
    ws_cat.freeze_panes = "A2"
    for i, c in enumerate(categories, start=2):
        ws_cat.cell(row=i, column=1, value=c.name)
        ws_cat.cell(row=i, column=2, value=c.description)

    # ── Sheet „Status (Referenz)" ─────────────────────────────────────────
    ws_st = wb.create_sheet("Status (Referenz)")
    cell = ws_st.cell(row=1, column=1, value="Bezeichnung")
    cell.font = header_font; cell.fill = header_fill
    ws_st.column_dimensions["A"].width = 30
    ws_st.freeze_panes = "A2"
    for i, (_, label) in enumerate(InventoryItem.Status.choices, start=2):
        ws_st.cell(row=i, column=1, value=label)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="inventar_vorlage.xlsx"'
    return response


def _parse_inventory_import_file(f):
    """xlsx parsen → (rows, errors). rows sind JSON-serialisierbare Dicts."""
    from collections import defaultdict
    import openpyxl

    if not f.name.lower().endswith(".xlsx"):
        return [], [{"row": 0, "messages": ["Bitte eine .xlsx-Datei verwenden."]}]

    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb["Import"] if "Import" in wb.sheetnames else wb.active
    raw_rows = list(ws.iter_rows(values_only=True))

    if not raw_rows:
        return [], [{"row": 0, "messages": ["Die Datei ist leer."]}]

    header = [str(h).strip().lower() if h is not None else "" for h in raw_rows[0]]
    data_rows = raw_rows[1:]

    cat_by_name: dict[str, list] = defaultdict(list)
    for c in InventoryCategory.objects.all():
        cat_by_name[c.name.strip().lower()].append(c)

    status_label_to_key = {label.lower(): key for key, label in InventoryItem.Status.choices}
    status_options = ", ".join(label for _, label in InventoryItem.Status.choices)

    parsed_rows = []
    errors = []

    for i, raw in enumerate(data_rows, start=2):
        if not any(v not in (None, "") for v in raw):
            continue
        row = {
            header[j]: str(v).strip() if v is not None else ""
            for j, v in enumerate(raw)
            if j < len(header) and header[j]
        }
        row_errors: list[str] = []

        name = row.get("bezeichnung", "")
        if not name:
            row_errors.append("Bezeichnung fehlt.")
        elif len(name) > 200:
            row_errors.append(f'Bezeichnung zu lang (max. 200 Zeichen): „{name[:50]}…".')

        cat_name = row.get("kategorie", "")
        cat_id = None
        if not cat_name:
            row_errors.append("Kategorie fehlt.")
        else:
            matches = cat_by_name.get(cat_name.lower(), [])
            if not matches:
                row_errors.append(f'Unbekannte Kategorie: „{cat_name}".')
            elif len(matches) > 1:
                row_errors.append(
                    f'Kategorie „{cat_name}" ist mehrdeutig ({len(matches)} Treffer).'
                )
            else:
                cat_id = matches[0].pk

        serial = row.get("seriennummer", "")
        if len(serial) > 100:
            row_errors.append("Seriennummer zu lang (max. 100 Zeichen).")

        status_raw = row.get("status", "")
        status_key = InventoryItem.Status.AVAILABLE
        if status_raw:
            key = status_label_to_key.get(status_raw.lower())
            if not key:
                row_errors.append(
                    f'Unbekannter Status: „{status_raw}" (erlaubt: {status_options}).'
                )
            else:
                status_key = key

        location = row.get("lagerort", "")
        if len(location) > 100:
            row_errors.append("Lagerort zu lang (max. 100 Zeichen).")

        parsed = {
            "row_num":       i,
            "name":          name,
            "category_id":   cat_id,
            "category_name": cat_name,
            "serial_number": serial,
            "status":        status_key,
            "status_label":  dict(InventoryItem.Status.choices).get(status_key, status_key),
            "location":      location,
            "notes":         row.get("notizen", ""),
            "errors":        row_errors,
        }
        parsed_rows.append(parsed)
        if row_errors:
            errors.append({"row": i, "messages": row_errors})

    return parsed_rows, errors


def _commit_inventory_import(rows):
    """Importiert die gültigen Zeilen."""
    count = 0
    for r in rows:
        if r.get("errors"):
            continue
        if not r["name"] or not r["category_id"]:
            continue
        InventoryItem.objects.create(
            name=r["name"],
            category_id=r["category_id"],
            serial_number=r.get("serial_number", "") or "",
            status=r.get("status") or InventoryItem.Status.AVAILABLE,
            location=r.get("location", "") or "",
            notes=r.get("notes", "") or "",
        )
        count += 1
    return count


@login_required
def inventory_import(request):
    """Excel-Import für Inventargegenstände (zweistufig mit Vorschau)."""
    _require_manage(request)

    if request.method == "POST" and "confirm" in request.POST:
        rows = request.session.pop("inventory_import_rows", [])
        if not rows:
            messages.error(request, "Keine Vorschau-Daten gefunden. Bitte erneut hochladen.")
            return redirect("inventory:inventory_import")
        created = _commit_inventory_import(rows)
        messages.success(request, f"{created} Inventargegenstand/-stände erfolgreich importiert.")
        return redirect("inventory:inventory_list")

    if request.method == "POST":
        form = InventoryImportForm(request.POST, request.FILES)
        if form.is_valid():
            rows, errors = _parse_inventory_import_file(request.FILES["file"])
            if not rows:
                messages.error(request, "Die Datei enthält keine Daten.")
                return render(request, "inventory/inventory_import.html", {"form": form})
            request.session["inventory_import_rows"] = rows
            valid_count = sum(1 for r in rows if not r["errors"])
            return render(request, "inventory/inventory_import_preview.html", {
                "rows":        rows,
                "errors":      errors,
                "valid_count": valid_count,
            })
    else:
        form = InventoryImportForm()

    return render(request, "inventory/inventory_import.html", {"form": form})


# ── Schnell-Ausgabe per Scan/Seriennummer ────────────────────────────────────

@login_required
def quick_issue(request):
    """Sucht einen Gegenstand per Seriennummer (Scan oder Eingabe) und leitet zur Ausgabe weiter."""
    _require_manage(request)
    query = (request.GET.get("q") or request.POST.get("q") or "").strip()
    matches = []
    if query:
        # 1) Exakter Treffer auf Seriennummer → direkt weiterleiten
        exact = list(
            InventoryItem.objects
            .select_related("category")
            .filter(serial_number__iexact=query)
        )
        if len(exact) == 1:
            item = exact[0]
            if item.status == InventoryItem.Status.AVAILABLE:
                return redirect("inventory:issuance_create", item_pk=item.pk)
            messages.warning(
                request,
                f'„{item.name}" ist aktuell „{item.get_status_display()}" und kann nicht ausgegeben werden.',
            )
            return redirect("inventory:item_detail", public_id=item.public_id)
        # 2) Fuzzy-Suche über Name + Seriennummer
        from django.db.models import Q
        matches = list(
            InventoryItem.objects
            .select_related("category")
            .filter(Q(name__icontains=query) | Q(serial_number__icontains=query))
            .order_by("status", "category__name", "name")[:50]
        )
    return render(request, "inventory/quick_issue.html", {
        "query": query,
        "matches": matches,
    })


# ── Kategorien ────────────────────────────────────────────────────────────────

@login_required
def category_list(request):
    """Listenansicht aller Inventarkategorien."""
    _require_manage(request)
    categories = InventoryCategory.objects.select_related("receipt_template").prefetch_related("items").all()
    return render(request, "inventory/category_list.html", {"categories": categories})


@login_required
def category_create(request):
    """Neue Inventarkategorie anlegen."""
    _require_manage(request)
    templates = ReceiptTemplate.objects.filter(is_active=True)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        icon = request.POST.get("icon", "bi-box").strip()
        description = request.POST.get("description", "").strip()
        template_id = request.POST.get("receipt_template") or None
        if not name:
            messages.error(request, "Bezeichnung darf nicht leer sein.")
        else:
            cat = InventoryCategory.objects.create(
                name=name,
                icon=icon or "bi-box",
                description=description,
                receipt_template_id=template_id,
            )
            messages.success(request, f'Kategorie „{cat.name}" wurde erstellt.')
            return redirect("inventory:category_list")
    return render(request, "inventory/category_form.html", {
        "templates": templates,
        "action": "Neue Kategorie",
    })


@login_required
def category_edit(request, public_id):
    """Bestehende Inventarkategorie bearbeiten."""
    _require_manage(request)
    category = get_object_or_404(InventoryCategory, public_id=public_id)
    templates = ReceiptTemplate.objects.filter(is_active=True)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        icon = request.POST.get("icon", "bi-box").strip()
        description = request.POST.get("description", "").strip()
        template_id = request.POST.get("receipt_template") or None
        if not name:
            messages.error(request, "Bezeichnung darf nicht leer sein.")
        else:
            category.name = name
            category.icon = icon or "bi-box"
            category.description = description
            category.receipt_template_id = template_id
            category.save()
            messages.success(request, f'Kategorie „{category.name}" wurde gespeichert.')
            return redirect("inventory:category_list")
    return render(request, "inventory/category_form.html", {
        "category": category,
        "templates": templates,
        "action": "Kategorie bearbeiten",
    })


@login_required
@require_POST
def category_delete(request, public_id):
    """Inventarkategorie löschen (nur wenn keine Gegenstände enthalten)."""
    _require_manage(request)
    category = get_object_or_404(InventoryCategory, public_id=public_id)
    if category.items.exists():
        messages.error(request, "Kategorie kann nicht gelöscht werden, da noch Gegenstände vorhanden sind.")
        return redirect("inventory:category_list")
    name = category.name
    category.delete()
    messages.success(request, f'Kategorie „{name}" wurde gelöscht.')
    return redirect("inventory:category_list")


# ── Gegenstände ───────────────────────────────────────────────────────────────

@login_required
def item_detail(request, public_id):
    """Detailansicht eines Inventargegenstands mit Ausgabehistorie."""
    _require_manage(request)
    item = get_object_or_404(
        InventoryItem.objects.select_related("category", "category__receipt_template"),
        public_id=public_id,
    )
    issuances = item.issuances.select_related("student", "issued_by", "returned_acknowledged_by").all()
    return render(request, "inventory/item_detail.html", {
        "item": item,
        "issuances": issuances,
        "status_choices": InventoryItem.Status.choices,
    })


@login_required
def item_create(request):
    """Neuen Inventargegenstand erfassen."""
    _require_manage(request)
    categories = InventoryCategory.objects.all()
    if request.method == "POST":
        category_id = request.POST.get("category")
        name = request.POST.get("name", "").strip()
        serial_number = request.POST.get("serial_number", "").strip()
        status = request.POST.get("status", InventoryItem.Status.AVAILABLE)
        location = request.POST.get("location", "").strip()
        notes = request.POST.get("notes", "").strip()
        if not name or not category_id:
            messages.error(request, "Bezeichnung und Kategorie sind Pflichtfelder.")
        else:
            item = InventoryItem.objects.create(
                category_id=category_id,
                name=name,
                serial_number=serial_number,
                status=status,
                location=location,
                notes=notes,
            )
            messages.success(request, f'Gegenstand „{item.name}" wurde erfasst.')
            return redirect("inventory:item_detail", public_id=item.public_id)
    return render(request, "inventory/item_form.html", {
        "categories": categories,
        "status_choices": InventoryItem.Status.choices,
        "action": "Neuer Gegenstand",
    })


@login_required
def item_edit(request, public_id):
    """Bestehenden Inventargegenstand bearbeiten."""
    _require_manage(request)
    item = get_object_or_404(InventoryItem, public_id=public_id)
    categories = InventoryCategory.objects.all()
    if request.method == "POST":
        category_id = request.POST.get("category")
        name = request.POST.get("name", "").strip()
        serial_number = request.POST.get("serial_number", "").strip()
        status = request.POST.get("status", item.status)
        location = request.POST.get("location", "").strip()
        notes = request.POST.get("notes", "").strip()
        if not name or not category_id:
            messages.error(request, "Bezeichnung und Kategorie sind Pflichtfelder.")
        else:
            item.category_id = category_id
            item.name = name
            item.serial_number = serial_number
            item.status = status
            item.location = location
            item.notes = notes
            item.save()
            messages.success(request, f'Gegenstand „{item.name}" wurde gespeichert.')
            return redirect("inventory:item_detail", public_id=item.public_id)
    return render(request, "inventory/item_form.html", {
        "item": item,
        "categories": categories,
        "status_choices": InventoryItem.Status.choices,
        "action": "Gegenstand bearbeiten",
    })


@login_required
@require_POST
def item_delete(request, public_id):
    """Inventargegenstand löschen (nur wenn nicht aktuell ausgegeben)."""
    _require_manage(request)
    item = get_object_or_404(InventoryItem, public_id=public_id)
    if item.issuances.filter(returned_at__isnull=True).exists():
        messages.error(request, "Gegenstand ist aktuell ausgegeben und kann nicht gelöscht werden.")
        return redirect("inventory:item_detail", public_id=public_id)
    name = item.name
    item.delete()
    messages.success(request, f'Gegenstand „{name}" wurde gelöscht.')
    return redirect("inventory:inventory_list")


# ── Ausgaben ──────────────────────────────────────────────────────────────────

@login_required
def issuance_create(request, item_public_id):
    """Gegenstand an eine Nachwuchskraft ausgeben und optional Quittung generieren."""
    _require_manage(request)
    item = get_object_or_404(
        InventoryItem.objects.select_related("category__receipt_template"),
        public_id=item_public_id,
    )
    students = Student.objects.order_by("last_name", "first_name")

    if request.method == "POST":
        student_id = request.POST.get("student")
        issued_at_raw = request.POST.get("issued_at", "").strip()
        notes = request.POST.get("notes", "").strip()

        student = Student.objects.filter(public_id=student_id).first()
        issued_at = None
        if issued_at_raw:
            issued_at = parse_datetime(issued_at_raw)
            if issued_at and issued_at.tzinfo is None:
                issued_at = timezone.make_aware(issued_at)

        if not student or not issued_at:
            messages.error(request, "Bitte Nachwuchskraft und Ausgabezeitpunkt angeben.")
            return render(request, "inventory/issuance_form.html", {
                "item": item,
                "students": students,
            })

        issuance = InventoryIssuance.objects.create(
            item=item,
            student=student,
            issued_by=request.user,
            issued_at=issued_at,
            notes=notes,
        )

        item.status = InventoryItem.Status.ISSUED
        item.save(update_fields=["status", "updated_at"])

        from services.notifications import notify_student_of_inventory_issuance
        notify_student_of_inventory_issuance(request, issuance)

        template = item.category.receipt_template
        if template and template.template_file:
            try:
                return _download_receipt(issuance, template)
            except Exception as exc:
                messages.warning(request, f'Ausgabe gespeichert, Quittungsgenerierung fehlgeschlagen: {exc}')
                return redirect("inventory:item_detail", public_id=item.public_id)
        else:
            messages.success(request, "Ausgabe gespeichert. Keine Quittungsvorlage für diese Kategorie hinterlegt.")
            return redirect("inventory:item_detail", public_id=item.public_id)

    return render(request, "inventory/issuance_form.html", {
        "item": item,
        "students": students,
    })


def _download_receipt(issuance, template) -> HttpResponse:
    """Generiert eine Ausgabequittung als DOCX aus der Vorlage und gibt sie als Download zurück."""
    from docxtpl import DocxTemplate, InlineImage
    from docx.shared import Mm
    from document.contexts import student_context, creator_context, meta_context

    # InlineImage benötigt das DocxTemplate-Objekt → Rendering hier statt in document.render
    doc = DocxTemplate(template.template_file.path)
    qr_buf = _make_qr_image(f"{QR_PREFIX}{issuance.pk}")
    ctx = {
        **student_context(issuance.student),
        **creator_context(issuance.issued_by),
        **meta_context(),
        "gegenstand_bezeichnung":  str(issuance.item),
        "gegenstand_seriennummer": issuance.item.serial_number,
        "gegenstand_kategorie":    issuance.item.category.name,
        "ausgabe_datum":           issuance.issued_at.strftime("%d.%m.%Y %H:%M"),
        "qr_code":                 InlineImage(doc, qr_buf, width=Mm(28)),
    }
    doc.render(ctx)
    buf = BytesIO()
    doc.save(buf)
    buf.seek(0)

    filename = f"Ausgabequittung_{issuance.item.name}_{issuance.student.last_name}_{issuance.issued_at.strftime('%Y%m%d')}.docx"
    filename_safe = filename.replace(" ", "_")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename_safe}"'
    return response


@login_required
def issuance_receipt_download(request, public_id):
    """Quittung für eine bestehende Ausgabe erneut herunterladen."""
    _require_manage(request)
    issuance = get_object_or_404(
        InventoryIssuance.objects.select_related(
            "item__category__receipt_template", "student", "issued_by"
        ),
        public_id=public_id,
    )
    template = issuance.item.category.receipt_template
    if not template or not template.template_file:
        messages.error(request, "Keine Quittungsvorlage für diese Kategorie hinterlegt.")
        return redirect("inventory:item_detail", public_id=issuance.item.public_id)
    try:
        return _download_receipt(issuance, template)
    except Exception as exc:
        messages.error(request, f"Quittungsgenerierung fehlgeschlagen: {exc}")
        return redirect("inventory:item_detail", public_id=issuance.item.public_id)


@login_required
@require_POST
def issuance_return(request, public_id):
    """Rückgabe eines ausgegebenen Gegenstands erfassen."""
    _require_manage(request)
    issuance = get_object_or_404(InventoryIssuance, public_id=public_id)
    if issuance.is_returned:
        messages.error(request, "Dieser Gegenstand wurde bereits zurückgegeben.")
        return redirect("inventory:item_detail", public_id=issuance.item.public_id)

    returned_at_raw = request.POST.get("returned_at", "").strip()
    returned_at = None
    if returned_at_raw:
        returned_at = parse_datetime(returned_at_raw)
        if returned_at and returned_at.tzinfo is None:
            returned_at = timezone.make_aware(returned_at)

    issuance.returned_at = returned_at or timezone.now()
    issuance.returned_acknowledged_by = request.user
    issuance.save(update_fields=["returned_at", "returned_acknowledged_by"])

    from services.notifications import notify_student_of_inventory_return
    notify_student_of_inventory_return(request, issuance)

    # Status auf "verfügbar" zurücksetzen, wenn keine weiteren offenen Ausgaben
    if not issuance.item.issuances.filter(returned_at__isnull=True).exists():
        issuance.item.status = InventoryItem.Status.AVAILABLE
        issuance.item.save(update_fields=["status", "updated_at"])

    messages.success(request, "Rückgabe wurde erfasst.")
    return redirect("inventory:item_detail", public_id=issuance.item.public_id)


@login_required
def scan_upload(request):
    """
    Nimmt eine oder mehrere eingescannte Quittungen (PDF) entgegen, liest den QR-Code,
    verknüpft jedes Dokument automatisch mit der Ausgabe und lädt es in Paperless hoch.
    """
    _require_manage(request)

    results = []
    if request.method == "POST":
        pdf_files = request.FILES.getlist("scan_pdf")
        if not pdf_files:
            messages.error(request, "Bitte mindestens eine PDF-Datei hochladen.")
        else:
            from services.validators import validate_pdf
            from django.core.exceptions import ValidationError
            for pdf_file in pdf_files:
                entry = {"filename": pdf_file.name, "success": False, "issuance": None, "doc_id": None, "error": None}
                try:
                    validate_pdf(pdf_file)
                except ValidationError as e:
                    entry["error"] = str(e.message)
                    results.append(entry)
                    continue
                pdf_bytes = pdf_file.read()
                qr_data = _read_qr_from_pdf(pdf_bytes)

                if not qr_data:
                    entry["error"] = "Kein QR-Code gefunden."
                    results.append(entry)
                    continue

                try:
                    issuance_pk = int(qr_data.removeprefix(QR_PREFIX))
                    issuance = InventoryIssuance.objects.select_related(
                        "item__category", "student", "issued_by"
                    ).get(pk=issuance_pk)
                except (ValueError, InventoryIssuance.DoesNotExist):
                    entry["error"] = f"QR-Code erkannt ({qr_data}), aber keine passende Ausgabe gefunden."
                    results.append(entry)
                    continue

                title = (
                    f"Ausgabequittung (unterschrieben) – {issuance.item} – "
                    f"{issuance.student} – {issuance.issued_at.strftime('%d.%m.%Y')}"
                )
                try:
                    doc_id = PaperlessService.upload_and_wait(
                        file_bytes=pdf_bytes,
                        title=title,
                        student_id=str(issuance.student.pk),
                        filename=pdf_file.name or "quittung.pdf",
                        mime_type="application/pdf",
                    )
                    if doc_id:
                        issuance.scanned_receipt_paperless_id = doc_id
                        issuance.save(update_fields=["scanned_receipt_paperless_id"])
                        entry.update({"success": True, "issuance": issuance, "doc_id": doc_id})
                    else:
                        entry["error"] = "Upload zu Paperless fehlgeschlagen."
                except Exception as exc:
                    entry["error"] = f"Fehler beim Upload: {exc}"
                results.append(entry)

            ok = sum(1 for r in results if r["success"])
            fail = len(results) - ok
            if ok:
                messages.success(request, f"{ok} Scan{'s' if ok != 1 else ''} erfolgreich verarbeitet." + (f" {fail} fehlgeschlagen." if fail else ""))
            elif fail:
                messages.error(request, f"Alle {fail} Scans konnten nicht verarbeitet werden.")

    return render(request, "inventory/scan_upload.html", {"results": results})


# ── Vorlagen ──────────────────────────────────────────────────────────────────

@login_required
def template_list(request):
    """Listenansicht aller Quittungsvorlagen mit verfügbaren Platzhaltern."""
    _require_manage(request)
    templates = ReceiptTemplate.objects.all()
    placeholders = [
        "student_vorname", "student_nachname", "student_id",
        "gegenstand_bezeichnung", "gegenstand_seriennummer", "gegenstand_kategorie",
        "ausgabe_datum", "ausgabe_von", "heute", "qr_code",
    ]
    return render(request, "inventory/template_list.html", {
        "templates": templates,
        "placeholders": placeholders,
    })


@login_required
def template_create(request):
    """Neue Quittungsvorlage hochladen."""
    _require_manage(request)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        is_active = request.POST.get("is_active") == "on"
        file = request.FILES.get("template_file")
        if not name or not file:
            messages.error(request, "Name und Datei sind Pflichtfelder.")
        else:
            from services.validators import validate_docx
            from django.core.exceptions import ValidationError
            try:
                validate_docx(file)
            except ValidationError as e:
                messages.error(request, str(e.message))
                return render(request, "inventory/template_form.html", {"action": "Neue Vorlage"})
            tmpl = ReceiptTemplate.objects.create(
                name=name,
                template_file=file,
                is_active=is_active,
            )
            messages.success(request, f'Vorlage „{tmpl.name}" wurde hochgeladen.')
            return redirect("inventory:template_list")
    return render(request, "inventory/template_form.html", {"action": "Neue Vorlage"})


@login_required
def template_edit(request, public_id):
    """Bestehende Quittungsvorlage bearbeiten."""
    _require_manage(request)
    tmpl = get_object_or_404(ReceiptTemplate, public_id=public_id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        is_active = request.POST.get("is_active") == "on"
        file = request.FILES.get("template_file")
        if not name:
            messages.error(request, "Name darf nicht leer sein.")
        else:
            tmpl.name = name
            tmpl.is_active = is_active
            if file:
                from services.validators import validate_docx
                from django.core.exceptions import ValidationError
                try:
                    validate_docx(file)
                except ValidationError as e:
                    messages.error(request, str(e.message))
                    return render(request, "inventory/template_form.html", {"tmpl": tmpl, "action": "Vorlage bearbeiten"})
                tmpl.template_file = file
            tmpl.save()
            messages.success(request, f'Vorlage „{tmpl.name}" wurde gespeichert.')
            return redirect("inventory:template_list")
    return render(request, "inventory/template_form.html", {
        "tmpl": tmpl,
        "action": "Vorlage bearbeiten",
    })


@login_required
@require_POST
def template_delete(request, public_id):
    """Quittungsvorlage löschen (nur wenn nicht mehr von Kategorien verwendet)."""
    _require_manage(request)
    tmpl = get_object_or_404(ReceiptTemplate, public_id=public_id)
    if tmpl.categories.exists():
        messages.error(request, "Vorlage wird noch von Kategorien verwendet und kann nicht gelöscht werden.")
        return redirect("inventory:template_list")
    name = tmpl.name
    tmpl.delete()
    messages.success(request, f'Vorlage „{name}" wurde gelöscht.')
    return redirect("inventory:template_list")
