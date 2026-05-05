# SPDX-License-Identifier: EUPL-1.2
# SPDX-FileCopyrightText: 2026 devNicoLax
"""Views für die Organisationsstruktur (Organisationseinheiten, Standorte, Kapazitätsplanung)."""
from datetime import date, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Competence, Location, OrganisationalUnit
from .forms import AdressForm, CompetenceForm, LocationForm, OrganisationalUnitForm, OrganisationalUnitImportForm


def _build_capacity_maps() -> tuple[dict[int, int], dict[int, int]]:
    """
    Gibt zwei Dicts zurück (jeweils nach OE-PK):
    - usage_map:   aktive Praktikumseinsätze heute (eigene + aller Untereinheiten)
    - max_map:     Summe der max_capacity (eigene + aller Untereinheiten)
    Benötigt nur 2 DB-Abfragen.
    """
    from course.models import InternshipAssignment

    today = date.today()

    all_units = list(OrganisationalUnit.objects.only('pk', 'parent_id', 'max_capacity'))
    unit_max: dict[int, int | None] = {u.pk: u.max_capacity for u in all_units}

    children_map: dict[int, list[int]] = {u.pk: [] for u in all_units}
    for u in all_units:
        if u.parent_id:
            children_map.setdefault(u.parent_id, []).append(u.pk)

    # Direkte Einsatzzahlen pro OE (heute aktiv)
    direct: dict[int, int] = {}
    for row in InternshipAssignment.objects.filter(
        start_date__lte=today, end_date__gte=today
    ).values('unit_id'):
        direct[row['unit_id']] = direct.get(row['unit_id'], 0) + 1

    usage_memo: dict[int, int] = {}
    max_memo: dict[int, int] = {}

    def calc_usage(pk: int) -> int:
        if pk in usage_memo:
            return usage_memo[pk]
        cnt = direct.get(pk, 0)
        for child_pk in children_map.get(pk, []):
            cnt += calc_usage(child_pk)
        usage_memo[pk] = cnt
        return cnt

    def calc_max(pk: int) -> int:
        if pk in max_memo:
            return max_memo[pk]
        total = unit_max.get(pk) or 0
        for child_pk in children_map.get(pk, []):
            total += calc_max(child_pk)
        max_memo[pk] = total
        return total

    all_pks = list(children_map.keys())
    return (
        {pk: calc_usage(pk) for pk in all_pks},
        {pk: calc_max(pk) for pk in all_pks},
    )


def _build_tree(units: list[OrganisationalUnit]) -> list[dict]:
    """Wandelt eine flache Liste in einen verschachtelten Baum um: [{"unit": obj, "children": [...]}, ...]"""
    by_id = {u.pk: {"unit": u, "children": []} for u in units}
    roots = []
    for node in by_id.values():
        parent_id = node["unit"].parent_id
        if parent_id and parent_id in by_id:
            by_id[parent_id]["children"].append(node)
        else:
            roots.append(node)
    return roots


@login_required
def unit_list(request):
    """Übersicht aller Organisationseinheiten als Baumstruktur."""
    from services.roles import is_training_director, is_training_office
    units = OrganisationalUnit.objects.select_related("parent").order_by(
        "unit_type", "parent__name", "name"
    )
    unit_list_flat = list(units)
    tree = _build_tree(unit_list_flat)
    total = len(unit_list_flat)
    active = sum(1 for u in unit_list_flat if u.is_active)
    return render(request, "organisation/unit_list.html", {
        "tree": tree,
        "total_count": total,
        "active_count": active,
        "inactive_count": total - active,
        "can_import": is_training_director(request.user) or is_training_office(request.user),
    })


def _get_coordinator_area_pks(user):
    """Gibt die PKs der OEs zurück, auf die eine Ausbildungskoordination Zugriff hat, oder None."""
    from services.roles import is_training_coordinator, get_chief_instructor
    if not (user.is_authenticated and is_training_coordinator(user)):
        return None
    from instructor.views import _get_coordination_area
    chief = get_chief_instructor(user)
    if not chief or not chief.coordination:
        return set()
    pks, _, _ = _get_coordination_area(chief.coordination)
    return set(pks)


@login_required
def unit_detail(request, public_id):
    """Detailansicht einer OE mit Einsätzen, Kapazitätskalender und Untereinheiten."""
    from django.core.exceptions import PermissionDenied
    coordinator_area_pks = _get_coordinator_area_pks(request.user)
    if coordinator_area_pks is not None and public_id not in coordinator_area_pks:
        raise PermissionDenied

    unit = get_object_or_404(OrganisationalUnit.objects.select_related("parent"), public_id=public_id)
    children = list(unit.children.order_by("unit_type", "name"))
    ancestors = unit.get_ancestors()

    usage_map, max_map = _build_capacity_maps()

    children_with_capacity = [
        (child, usage_map.get(child.public_id, 0), max_map.get(child.public_id, 0))
        for child in children
    ]

    # Alle PKs der Untereinheiten (inkl. eigener) sammeln
    all_units = list(OrganisationalUnit.objects.only('public_id', 'parent_id'))
    children_map: dict[int, list[int]] = {u.public_id: [] for u in all_units}
    for u in all_units:
        if u.parent_id:
            children_map.setdefault(u.parent_id, []).append(u.public_id)

    def collect_pks(root_pk: int) -> list[int]:
        result = [root_pk]
        for child_pk in children_map.get(root_pk, []):
            result.extend(collect_pks(child_pk))
        return result

    descendant_pks = collect_pks(unit.public_id)

    from course.models import InternshipAssignment
    all_assignments = (
        InternshipAssignment.objects
        .filter(unit_id__in=descendant_pks)
        .select_related('student', 'unit', 'schedule_block')
        .order_by('student__last_name', 'student__first_name')
    )

    today = date.today()
    current_assignments = [a for a in all_assignments if a.start_date <= today <= a.end_date]
    past_assignments = sorted([a for a in all_assignments if a.end_date < today], key=lambda a: a.end_date, reverse=True)
    future_assignments = sorted([a for a in all_assignments if a.start_date > today], key=lambda a: a.start_date)

    # Wöchentliche Kapazitätsaufschlüsselung für den Auslastungskalender (26 Wochen)
    week_base = today - timedelta(days=today.weekday())  # Monday of current week
    effective_max = max_map.get(unit.public_id, 0) or None  # aggregate max (self + descendants)
    capacity_weeks = []
    for i in range(26):
        ws = week_base + timedelta(weeks=i)
        we = ws + timedelta(days=6)
        week_items = [a for a in all_assignments if a.start_date <= we and a.end_date >= ws]
        count = len(week_items)
        if effective_max:
            if count == 0:
                status = 'free'
            elif count >= effective_max:
                status = 'danger'
            elif count / effective_max >= 0.75:
                status = 'warning'
            else:
                status = 'success'
        else:
            status = 'has_count' if count > 0 else 'free'
        capacity_weeks.append({
            'week_start': ws,
            'week_end': we,
            'count': count,
            'max': effective_max,
            'status': status,
            'pct': min(int(count / effective_max * 100), 100) if effective_max else None,
            'assignments': week_items,
        })

    own_locations = list(unit.locations.all())
    all_locations = list(unit.get_all_locations())
    inherited_locations = [loc for loc in all_locations if loc not in own_locations]

    return render(request, "organisation/unit_detail.html", {
        "unit": unit,
        "children": children,
        "children_with_capacity": children_with_capacity,
        "ancestors": ancestors,
        "unit_used": usage_map.get(unit.public_id, 0),
        "unit_total_max": max_map.get(unit.public_id, 0),
        "current_assignments": current_assignments,
        "past_assignments": past_assignments,
        "future_assignments": future_assignments,
        "coordinator_area_pks": coordinator_area_pks,
        "own_locations": own_locations,
        "inherited_locations": inherited_locations,
        "capacity_weeks": capacity_weeks,
        "effective_max": effective_max,
    })


@login_required
def unit_create(request):
    """Neue Organisationseinheit anlegen (Eltern-OE optional per Query-Parameter vorauswählbar)."""
    # Eltern-OE per Query-Parameter vorauswählen, z.B. ?parent=3
    initial = {}
    parent_pk = request.GET.get("parent")
    if parent_pk:
        parent = OrganisationalUnit.objects.filter(pk=parent_pk).first()
        if parent:
            initial["parent"] = parent

    form = OrganisationalUnitForm(request.POST or None, initial=initial)
    if form.is_valid():
        unit = form.save()
        messages.success(request, f'„{unit.name}" wurde erfolgreich angelegt.')
        return redirect("organisation:unit_detail", public_id=unit.public_id)
    return render(request, "organisation/unit_form.html", {"form": form, "action": "Anlegen"})


@login_required
def unit_edit(request, public_id):
    """Bestehende Organisationseinheit bearbeiten."""
    from django.core.exceptions import PermissionDenied
    coordinator_area_pks = _get_coordinator_area_pks(request.user)
    if coordinator_area_pks is not None and public_id not in coordinator_area_pks:
        raise PermissionDenied

    unit = get_object_or_404(OrganisationalUnit, public_id=public_id)
    form = OrganisationalUnitForm(request.POST or None, instance=unit)
    if form.is_valid():
        form.save()
        messages.success(request, f'„{unit.name}" wurde erfolgreich gespeichert.')
        return redirect("organisation:unit_detail", public_id=unit.public_id)
    return render(request, "organisation/unit_form.html", {
        "form": form,
        "action": "Bearbeiten",
        "unit": unit,
    })


@login_required
def location_list(request):
    """Listenansicht aller Standorte."""
    locations = Location.objects.select_related("address").order_by("name")
    return render(request, "organisation/location_list.html", {"locations": locations})


@login_required
def location_create(request):
    """Neuen Standort mit Adresse anlegen."""
    location_form = LocationForm(request.POST or None)
    address_form = AdressForm(request.POST or None)
    if location_form.is_valid() and address_form.is_valid():
        address = address_form.save()
        location = location_form.save(commit=False)
        location.address = address
        location.save()
        messages.success(request, f'Standort „{location.name}" wurde angelegt.')
        return redirect("organisation:location_list")
    return render(request, "organisation/location_form.html", {
        "location_form": location_form,
        "address_form": address_form,
        "action": "Anlegen",
    })


@login_required
def location_edit(request, public_id):
    """Bestehenden Standort bearbeiten."""
    location = get_object_or_404(Location.objects.select_related("address"), public_id=public_id)
    location_form = LocationForm(request.POST or None, instance=location)
    address_form = AdressForm(request.POST or None, instance=location.address)
    if location_form.is_valid() and address_form.is_valid():
        address_form.save()
        location_form.save()
        messages.success(request, f'Standort „{location.name}" wurde gespeichert.')
        return redirect("organisation:location_list")
    return render(request, "organisation/location_form.html", {
        "location_form": location_form,
        "address_form": address_form,
        "action": "Bearbeiten",
        "location": location,
    })


@login_required
def location_delete(request, public_id):
    """Standort inkl. Adresse löschen (mit Bestätigungsseite)."""
    location = get_object_or_404(Location.objects.select_related("address"), public_id=public_id)
    if request.method == "POST":
        name = location.name
        address = location.address
        location.delete()
        if address:
            address.delete()
        messages.success(request, f'Standort „{name}" wurde gelöscht.')
        return redirect("organisation:location_list")
    return render(request, "organisation/location_confirm_delete.html", {"location": location})


# ── Excel-Import ────────────────────────────────────────────────────────────


def _check_import_permission(user):
    from django.core.exceptions import PermissionDenied
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(user) or is_training_office(user)):
        raise PermissionDenied


@login_required
def unit_import_template(request):
    """Erzeugt eine Excel-Vorlage mit Kopfzeile, Beispielzeile und Referenzlisten."""
    _check_import_permission(request.user)

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from django.http import HttpResponse

    wb = openpyxl.Workbook()

    # ── Sheet „Import" ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Import"
    headers = [
        "Name", "Bezeichnung", "Art", "Übergeordnete Einheit",
        "Standorte", "Kompetenzen", "Max. Kapazität", "Notizen", "Aktiv",
    ]
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [16, 36, 18, 28, 40, 40, 14, 30, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Beispielzeile (kann gelöscht werden)
    example = ["REF-IT", "Referat IT", "Referat", "ABT-Z",
               "Hauptstandort; Außenstelle Süd", "IT-Sicherheit",
               20, "Beispielzeile – kann gelöscht werden", "ja"]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(italic=True, color="888888")

    # Drop-down für „Art"
    art_values = ",".join(label for _, label in OrganisationalUnit.UNIT_TYPES)
    art_dv = DataValidation(type="list", formula1=f'"{art_values}"', allow_blank=False)
    art_dv.error = "Bitte einen der vorgegebenen Werte wählen."
    art_dv.errorTitle = "Ungültige Art"
    art_dv.prompt = "Behörde, Abteilung, Referatsgruppe, Referat oder Sachgebiet"
    art_dv.promptTitle = "Art der Organisationseinheit"
    ws.add_data_validation(art_dv)
    art_dv.add("C2:C1000")

    # Drop-down für „Aktiv"
    aktiv_dv = DataValidation(type="list", formula1='"ja,nein"', allow_blank=True)
    ws.add_data_validation(aktiv_dv)
    aktiv_dv.add("I2:I1000")

    # ── Sheet „Standorte (Referenz)" ────────────────────────────────────
    ws_loc = wb.create_sheet("Standorte (Referenz)")
    ws_loc.cell(row=1, column=1, value="Name").font = Font(bold=True)
    ws_loc.cell(row=1, column=1).fill = header_fill
    ws_loc.cell(row=1, column=1).font = header_font
    ws_loc.column_dimensions["A"].width = 60
    ws_loc.freeze_panes = "A2"
    for i, loc in enumerate(Location.objects.order_by("name"), start=2):
        ws_loc.cell(row=i, column=1, value=loc.name)

    # ── Sheet „Kompetenzen (Referenz)" ──────────────────────────────────
    ws_comp = wb.create_sheet("Kompetenzen (Referenz)")
    ws_comp.cell(row=1, column=1, value="Name").font = header_font
    ws_comp.cell(row=1, column=1).fill = header_fill
    ws_comp.column_dimensions["A"].width = 60
    ws_comp.freeze_panes = "A2"
    for i, c in enumerate(Competence.objects.order_by("name"), start=2):
        ws_comp.cell(row=i, column=1, value=c.name)

    # ── Sheet „Bestehende OEs (Referenz)" ───────────────────────────────
    ws_oe = wb.create_sheet("Bestehende OEs (Referenz)")
    oe_headers = ["Name", "Bezeichnung", "Art", "Pfad"]
    for col_idx, h in enumerate(oe_headers, start=1):
        cell = ws_oe.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for col_letter, w in zip("ABCD", (20, 40, 18, 60)):
        ws_oe.column_dimensions[col_letter].width = w
    ws_oe.freeze_panes = "A2"
    type_label_map = dict(OrganisationalUnit.UNIT_TYPES)
    for i, u in enumerate(
        OrganisationalUnit.objects.select_related("parent").order_by("unit_type", "name"),
        start=2,
    ):
        ws_oe.cell(row=i, column=1, value=u.name)
        ws_oe.cell(row=i, column=2, value=u.label)
        ws_oe.cell(row=i, column=3, value=type_label_map.get(u.unit_type, u.unit_type))
        ws_oe.cell(row=i, column=4, value=u.get_full_path())

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="organisationseinheiten_vorlage.xlsx"'
    return response


def _parse_unit_import_file(f):
    """xlsx parsen → (rows, errors). rows sind JSON-serialisierbare Dicts."""
    from collections import defaultdict
    import openpyxl

    if not f.name.lower().endswith(".xlsx"):
        return [], [{"row": 0, "messages": ["Bitte eine .xlsx-Datei verwenden."]}]

    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb["Import"] if "Import" in wb.sheetnames else wb.active
    raw_rows = list(ws.iter_rows(values_only=True))

    if not raw_rows:
        return [], [{"row": 0, "messages": ["Die Datei ist leer."]}]

    header = [str(h).strip().lower() if h is not None else "" for h in raw_rows[0]]
    data_rows = raw_rows[1:]

    type_label_to_key = {label.lower(): key for key, label in OrganisationalUnit.UNIT_TYPES}
    type_label_options = ", ".join(label for _, label in OrganisationalUnit.UNIT_TYPES)

    locations = {loc.name.strip().lower(): loc for loc in Location.objects.all()}
    competences = {c.name.strip().lower(): c for c in Competence.objects.all()}

    db_units_by_name: dict[str, list] = defaultdict(list)
    for u in OrganisationalUnit.objects.all():
        db_units_by_name[u.name.strip().lower()].append(u)

    parsed_rows = []
    file_name_rows: dict[str, list[int]] = defaultdict(list)

    for i, raw in enumerate(data_rows, start=2):
        if not any(v not in (None, "") for v in raw):
            continue

        row = {
            header[j]: str(v).strip() if v is not None else ""
            for j, v in enumerate(raw)
            if j < len(header) and header[j]
        }
        row_errors: list[str] = []

        name_val = row.get("name", "")
        if not name_val:
            row_errors.append("Name fehlt.")
        elif len(name_val) > 30:
            row_errors.append(f'Name zu lang (max. 30 Zeichen): „{name_val}".')

        label_val = row.get("bezeichnung", "")
        if not label_val:
            row_errors.append("Bezeichnung fehlt.")

        art_raw = row.get("art", "")
        unit_type = type_label_to_key.get(art_raw.lower())
        if not unit_type:
            row_errors.append(
                f'Unbekannte Art: „{art_raw}" (erlaubt: {type_label_options}).'
            )

        parent_name = row.get("übergeordnete einheit", "")
        parent_id = None
        parent_pending = None
        if parent_name:
            db_matches = db_units_by_name.get(parent_name.lower(), [])
            if len(db_matches) > 1:
                row_errors.append(
                    f'Übergeordnete Einheit „{parent_name}" ist im System mehrdeutig '
                    f'({len(db_matches)} Treffer).'
                )
            elif len(db_matches) == 1:
                parent_id = db_matches[0].pk
            else:
                parent_pending = parent_name.lower()

        location_ids: list[int] = []
        for entry in row.get("standorte", "").split(";"):
            key = entry.strip().lower()
            if not key:
                continue
            loc = locations.get(key)
            if loc:
                location_ids.append(loc.pk)
            else:
                row_errors.append(f'Unbekannter Standort: „{entry.strip()}".')

        competence_ids: list[int] = []
        for entry in row.get("kompetenzen", "").split(";"):
            key = entry.strip().lower()
            if not key:
                continue
            comp = competences.get(key)
            if comp:
                competence_ids.append(comp.pk)
            else:
                row_errors.append(f'Unbekannte Kompetenz: „{entry.strip()}".')

        cap_raw = row.get("max. kapazität", "")
        max_capacity = None
        if cap_raw:
            try:
                max_capacity = int(float(cap_raw))
                if max_capacity < 0:
                    row_errors.append("Max. Kapazität darf nicht negativ sein.")
                    max_capacity = None
            except ValueError:
                row_errors.append(f'Max. Kapazität ist keine Zahl: „{cap_raw}".')

        aktiv_raw = row.get("aktiv", "").lower()
        if aktiv_raw in ("", "ja", "yes", "true", "1"):
            is_active = True
        elif aktiv_raw in ("nein", "no", "false", "0"):
            is_active = False
        else:
            row_errors.append(f'Aktiv: „{aktiv_raw}" – erlaubt sind „ja" oder „nein".')
            is_active = True

        parsed_rows.append({
            "row_num": i,
            "name": name_val,
            "label": label_val,
            "unit_type": unit_type,
            "unit_type_label": art_raw,
            "parent_name": parent_name,
            "parent_id": parent_id,
            "parent_pending": parent_pending,
            "location_ids": location_ids,
            "competence_ids": competence_ids,
            "max_capacity": max_capacity,
            "notes": row.get("notizen", ""),
            "is_active": is_active,
            "errors": row_errors,
        })
        if name_val:
            file_name_rows[name_val.strip().lower()].append(i)

    # Zweiter Durchgang: noch nicht aufgelöste Eltern gegen Datei prüfen
    errors = []
    for r in parsed_rows:
        if r["parent_pending"] and not r["parent_id"]:
            file_matches = file_name_rows.get(r["parent_pending"], [])
            if not file_matches:
                r["errors"].append(
                    f'Übergeordnete Einheit „{r["parent_name"]}" wurde weder im System '
                    'noch in der Datei gefunden.'
                )
            elif len(file_matches) > 1:
                r["errors"].append(
                    f'Übergeordnete Einheit „{r["parent_name"]}" ist in der Datei mehrdeutig '
                    f'(Zeilen: {", ".join(map(str, file_matches))}).'
                )
        if r["errors"]:
            errors.append({"row": r["row_num"], "messages": r["errors"]})

    return parsed_rows, errors


def _commit_unit_import(rows):
    """Importiert die gültigen Zeilen, sortiert nach Eltern-Abhängigkeit."""
    valid = [r for r in rows if not r["errors"]]
    if not valid:
        return 0

    name_to_pk: dict[str, int] = {}
    pending = list(valid)
    created = 0
    max_iter = len(pending) + 1

    while pending and max_iter > 0:
        next_pending = []
        for r in pending:
            parent_id = r.get("parent_id")
            parent_pending = r.get("parent_pending")
            if parent_pending and not parent_id:
                if parent_pending in name_to_pk:
                    parent_id = name_to_pk[parent_pending]
                else:
                    next_pending.append(r)
                    continue

            unit = OrganisationalUnit.objects.create(
                name=r["name"],
                label=r["label"],
                unit_type=r["unit_type"],
                parent_id=parent_id,
                max_capacity=r.get("max_capacity"),
                notes=r.get("notes") or "",
                is_active=r.get("is_active", True),
            )
            if r.get("location_ids"):
                unit.locations.set(r["location_ids"])
            if r.get("competence_ids"):
                unit.competences.set(r["competence_ids"])

            name_to_pk[r["name"].strip().lower()] = unit.pk
            created += 1

        if len(next_pending) == len(pending):
            break  # nicht auflösbare Abhängigkeit – Schleife verlassen
        pending = next_pending
        max_iter -= 1

    return created


@login_required
def unit_import(request):
    """Excel-Import für Organisationseinheiten (zweistufig mit Vorschau)."""
    _check_import_permission(request.user)

    if request.method == "POST" and "confirm" in request.POST:
        rows = request.session.pop("unit_import_rows", [])
        if not rows:
            messages.error(request, "Keine Vorschau-Daten gefunden. Bitte erneut hochladen.")
            return redirect("organisation:unit_import")
        created = _commit_unit_import(rows)
        messages.success(request, f"{created} Organisationseinheit(en) erfolgreich importiert.")
        return redirect("organisation:unit_list")

    if request.method == "POST":
        form = OrganisationalUnitImportForm(request.POST, request.FILES)
        if form.is_valid():
            rows, errors = _parse_unit_import_file(request.FILES["file"])
            if not rows:
                messages.error(request, "Die Datei enthält keine Daten.")
                return render(request, "organisation/unit_import.html", {"form": form})
            request.session["unit_import_rows"] = rows
            valid_count = sum(1 for r in rows if not r["errors"])
            return render(request, "organisation/unit_import_preview.html", {
                "rows": rows,
                "errors": errors,
                "valid_count": valid_count,
            })
    else:
        form = OrganisationalUnitImportForm()

    return render(request, "organisation/unit_import.html", {"form": form})


@login_required
def unit_delete(request, public_id):
    """Organisationseinheit löschen (mit Bestätigungsseite, Fehler bei vorhandenen Untereinheiten)."""
    unit = get_object_or_404(OrganisationalUnit, public_id=public_id)
    parent_pk = unit.parent_id
    if request.method == "POST":
        name = unit.name
        try:
            unit.delete()
            messages.success(request, f'„{name}" wurde gelöscht.')
        except Exception:
            messages.error(
                request,
                f'„{name}" kann nicht gelöscht werden, da noch untergeordnete Einheiten existieren.'
            )
            return redirect("organisation:unit_detail", public_id=public_id)
        if parent_pk:
            return redirect("organisation:unit_detail", public_id=parent_pk)
        return redirect("organisation:unit_list")
    return render(request, "organisation/unit_confirm_delete.html", {"unit": unit})


# ── Kompetenzen-Verwaltung ──────────────────────────────────────────────────

@login_required
def competence_list(request):
    """Übersicht aller Kompetenzen mit Verwendungs-Zählern."""
    from django.db.models import Count
    qs = (
        Competence.objects
        .annotate(
            unit_count=Count('units', distinct=True),
            criterion_count=Count('criterion_weights', distinct=True),
        )
        .order_by('name')
    )
    return render(request, "organisation/competence_list.html", {"competences": qs})


@login_required
def competence_create(request):
    if request.method == "POST":
        form = CompetenceForm(request.POST)
        if form.is_valid():
            comp = form.save()
            messages.success(request, f'Kompetenz „{comp.name}" wurde angelegt.')
            return redirect("organisation:competence_list")
    else:
        form = CompetenceForm()
    return render(request, "organisation/competence_form.html", {
        "form": form, "action": "Anlegen",
    })


@login_required
def competence_edit(request, public_id):
    comp = get_object_or_404(Competence, public_id=public_id)
    if request.method == "POST":
        form = CompetenceForm(request.POST, instance=comp)
        if form.is_valid():
            form.save()
            messages.success(request, f'Kompetenz „{comp.name}" wurde aktualisiert.')
            return redirect("organisation:competence_list")
    else:
        form = CompetenceForm(instance=comp)
    return render(request, "organisation/competence_form.html", {
        "form": form, "action": "Bearbeiten", "competence": comp,
    })


@login_required
def competence_delete(request, public_id):
    comp = get_object_or_404(Competence, public_id=public_id)
    if request.method == "POST":
        name = comp.name
        try:
            comp.delete()
            messages.success(request, f'„{name}" wurde gelöscht.')
        except Exception as exc:
            messages.error(request, f'„{name}" konnte nicht gelöscht werden: {exc}')
            return redirect("organisation:competence_list")
        return redirect("organisation:competence_list")
    return render(request, "organisation/competence_confirm_delete.html", {"competence": comp})
