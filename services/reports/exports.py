# SPDX-License-Identifier: EUPL-1.2
# SPDX-FileCopyrightText: 2026 devNicoLax
"""Excel- und CSV-Export für Reports (Spalten-typ-aware, mit optionaler Summenzeile)."""
from __future__ import annotations

import csv
import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import BaseReport, Column


_HEADER_FILL = PatternFill(start_color='FF0077B6', end_color='FF0077B6', fill_type='solid')
_HEADER_FONT = Font(bold=True, color='FFFFFFFF')
_TOTAL_FONT  = Font(bold=True)
_TOTAL_FILL  = PatternFill(start_color='FFEEF0F3', end_color='FFEEF0F3', fill_type='solid')


def _format_for_excel(value, col: Column):
    if value is None or value == '':
        return None
    if col.type == 'date':
        if isinstance(value, str):
            try:
                return date.fromisoformat(value)
            except ValueError:
                return value
        return value
    if col.type == 'datetime':
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value)
            except ValueError:
                return value
        return value
    if col.type == 'pct':
        try:
            return float(value) / 100.0
        except (ValueError, TypeError):
            return value
    if col.type == 'int':
        try:
            return int(value)
        except (ValueError, TypeError):
            return value
    if col.type == 'float':
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    return str(value)


def _column_number_format(col: Column) -> str | None:
    if col.type == 'date':
        return 'DD.MM.YYYY'
    if col.type == 'datetime':
        return 'DD.MM.YYYY HH:MM'
    if col.type == 'pct':
        return '0.0%'
    if col.type == 'int':
        return '0'
    if col.type == 'float':
        return '0.00'
    return None


def report_to_xlsx(report: BaseReport, rows: list[dict], visible_cols: list[Column]) -> bytes:
    """Baut eine XLSX-Datei aus Report-Daten."""
    wb = Workbook()
    ws = wb.active
    ws.title = report.name[:31] or 'Report'

    # Header
    for col_idx, col in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Daten
    totals: dict[int, float] = {}
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(visible_cols, start=1):
            value = _format_for_excel(row.get(col.key, col.default), col)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            fmt = _column_number_format(col)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal=col.align, vertical='center')
            if col.total and isinstance(value, (int, float)):
                totals[col_idx] = totals.get(col_idx, 0.0) + value

    # Summenzeile (falls mind. eine Spalte aufsummiert)
    if totals:
        total_row = len(rows) + 2
        label_cell = ws.cell(row=total_row, column=1, value='Summe')
        label_cell.font = _TOTAL_FONT
        label_cell.fill = _TOTAL_FILL
        for col_idx, value in totals.items():
            col = visible_cols[col_idx - 1]
            cell = ws.cell(row=total_row, column=col_idx, value=value)
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL
            fmt = _column_number_format(col)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal=col.align)

    # Spaltenbreiten näherungsweise
    for col_idx, col in enumerate(visible_cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, len(col.label) + 6))
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_to_csv(rows: list[dict], visible_cols: list[Column]) -> bytes:
    """Baut eine CSV-Datei (UTF-8 mit BOM, damit Excel sie korrekt öffnet)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow([c.label for c in visible_cols])
    for row in rows:
        writer.writerow([_csv_format(row.get(c.key, c.default), c) for c in visible_cols])
    return ('\ufeff' + buf.getvalue()).encode('utf-8')


def _csv_format(value, col: Column):
    if value is None or value == '':
        return ''
    if col.type == 'date' and hasattr(value, 'strftime'):
        return value.strftime('%d.%m.%Y')
    if col.type == 'datetime' and hasattr(value, 'strftime'):
        return value.strftime('%d.%m.%Y %H:%M')
    if col.type == 'pct':
        try:
            return f'{float(value):.1f}%'.replace('.', ',')
        except (ValueError, TypeError):
            return str(value)
    if col.type == 'float':
        try:
            return f'{float(value):.2f}'.replace('.', ',')
        except (ValueError, TypeError):
            return str(value)
    return str(value)
