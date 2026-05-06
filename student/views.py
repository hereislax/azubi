"""
Views für die Nachwuchskräfte-Verwaltung (Student-App).

Enthält Listen-, Detail-, Bearbeitungs-, Import- und Export-Views sowie
Noten, Checklisten, Kontakteinträge und Notizen.
"""
# SPDX-License-Identifier: EUPL-1.2
# SPDX-FileCopyrightText: 2026 devNicoLax
import csv
import io
from collections import Counter
from datetime import date, datetime

from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST

from course.models import ScheduleBlock
from services.paperless import PaperlessService
from .forms import StudentForm, AdressForm, GradeForm, StudentImportForm, StudentDocumentTemplateFieldFormSet
from .models import (
    Student, StudentFieldDefinition, StudentFieldValue, Status, Grade,
    StudentDocumentTemplate, StudentDocumentTemplateField, ContactEntry,
    ChecklistTemplate, StudentChecklist, StudentChecklistItem,
)
from .anonymization import (
    anonymize_student,
    get_students_due_for_anonymization,
    get_students_approaching_anonymization,
)

# Status-Beschreibungen, die als „aktiv" gelten
ACTIVE_STATUS_NAMES = {'aktiv'}


def _is_active(student):
    """True wenn der Status des Studierenden als aktiv gilt."""
    if student.status is None:
        return True  # kein Status = noch kein Abschluss → aktiv
    return student.status.description.strip().lower() in ACTIVE_STATUS_NAMES

@login_required
def student_list(request):
    from services.roles import (
        is_training_coordinator, get_chief_instructor,
        is_dormitory_management, get_dormitory_management_profile, is_training_responsible,
        is_training_office, is_training_director, get_training_office_profile,
    )
    from instructor.views import _get_coordination_area

    all_statuses = Status.objects.all()
    status_filter = request.GET.get('status', 'aktiv')

    qs = Student.objects.select_related('gender', 'course', 'status').order_by('last_name', 'first_name')

    # Ausbildungsreferat: Berufsbilder-Filter (wenn Toggle nicht auf "alle" steht)
    is_referat = request.user.is_authenticated and is_training_office(request.user)
    is_leitung = request.user.is_authenticated and is_training_director(request.user)
    if is_referat and not is_leitung and not request.session.get('training_office_show_all', False):
        training_office_profile = get_training_office_profile(request.user)
        if training_office_profile:
            jp_pks = list(training_office_profile.job_profiles.values_list('pk', flat=True))
            if jp_pks:
                qs = qs.filter(course__job_profile__pk__in=jp_pks)

    # Ausbildungskoordination: sieht nur zugewiesene Nachwuchskräfte
    restricted = False
    from django.db.models import Q
    is_koord = request.user.is_authenticated and is_training_coordinator(request.user)
    is_ps_role = request.user.is_authenticated and is_training_responsible(request.user)

    if is_koord:
        restricted = True
        chief = get_chief_instructor(request.user)
        koord_pks = set()
        if chief and chief.coordination:
            from course.models import InternshipAssignment
            from datetime import timedelta
            descendant_pks, _, _ = _get_coordination_area(chief.coordination)
            cutoff = date.today() - timedelta(days=14)
            koord_pks = set(
                InternshipAssignment.objects.filter(
                    unit_id__in=descendant_pks,
                    end_date__gte=cutoff,
                ).values_list('student_id', flat=True)
            )

        if is_ps_role:
            from student.models import TrainingResponsibleAccess
            ps_pks = set(
                TrainingResponsibleAccess.objects.filter(
                    user=request.user,
                ).values_list('student_id', flat=True)
            )
            allowed_pks = koord_pks | ps_pks
        else:
            allowed_pks = koord_pks

        qs = qs.filter(pk__in=allowed_pks)

    elif request.user.is_authenticated and is_dormitory_management(request.user):
        restricted = True
        hv_profile = get_dormitory_management_profile(request.user)
        if hv_profile:
            from dormitory.models import RoomAssignment as DormRoomAssignment
            from django.db.models import Q
            from datetime import timedelta
            cutoff = date.today() + timedelta(days=90)
            student_pks = DormRoomAssignment.objects.filter(
                room__dormitory=hv_profile.dormitory,
                start_date__lte=cutoff,
            ).filter(
                Q(end_date__isnull=True) | Q(end_date__gte=date.today())
            ).values_list('student_id', flat=True)
            qs = qs.filter(pk__in=student_pks)
        else:
            qs = qs.none()

    elif is_ps_role:
        # Ausbildungsverantwortliche: nur freigegebene Nachwuchskräfte
        restricted = True
        from student.models import TrainingResponsibleAccess
        student_pks = TrainingResponsibleAccess.objects.filter(
            user=request.user,
        ).values_list('student_id', flat=True)
        qs = qs.filter(pk__in=student_pks)

    if status_filter == 'alle':
        students = qs
    elif status_filter == 'kein':
        students = qs.filter(status__isnull=True)
    else:
        students = qs.filter(status__description__iexact=status_filter)

    for s in students:
        s.computed_active = _is_active(s)

    return render(request, 'student/student_list.html', {
        'students': students,
        'all_statuses': all_statuses,
        'status_filter': status_filter,
        'restricted': restricted,
    })

@login_required
def student_detail(request, pk):
    from services.roles import (
        is_training_coordinator, get_chief_instructor,
        is_dormitory_management, get_dormitory_management_profile,
        is_travel_expense_office, is_training_responsible,
    )
    from instructor.views import _get_coordination_area

    today = date.today()
    student = get_object_or_404(
        Student.objects.select_related('gender', 'course', 'employment', 'status'),
        pk=pk,
    )

    # Rollen-Flags für diese Ansicht
    is_hv_view = False
    is_rk_view = False
    is_ps_view = False
    koord_unit_pks = None  # Koordination: filtert Praktika auf eigene Einheiten

    # Ausbildungskoordination: nur zugewiesene Nachwuchskräfte, eingeschränkte Daten
    restricted_view = False
    can_edit = True
    is_koord = request.user.is_authenticated and is_training_coordinator(request.user)
    is_ps_role = request.user.is_authenticated and is_training_responsible(request.user)

    if is_koord:
        can_edit = False
        chief = get_chief_instructor(request.user)

        # Ausbildungsverantwortliche-Zugang vorab prüfen (kombinierte Rolle)
        ps_grants_access = False
        if is_ps_role:
            from student.models import TrainingResponsibleAccess
            ps_grants_access = TrainingResponsibleAccess.objects.filter(
                user=request.user, student=student
            ).exists()

        if chief and chief.coordination:
            restricted_view = True
            from course.models import InternshipAssignment
            from datetime import timedelta
            descendant_pks, _, _ = _get_coordination_area(chief.coordination)
            koord_unit_pks = descendant_pks
            cutoff = date.today() - timedelta(days=14)
            koord_has_access = InternshipAssignment.objects.filter(
                student=student, unit_id__in=descendant_pks, end_date__gte=cutoff
            ).exists()
            if not koord_has_access and not ps_grants_access:
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied
        else:
            if not ps_grants_access:
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied

        # Wenn die Personalstelle Zugriff gewährt, erweiterte Ansicht mit vollständigen PS-Tabs anzeigen
        if ps_grants_access:
            is_ps_view = True
            restricted_view = False

    elif request.user.is_authenticated and is_dormitory_management(request.user):
        is_hv_view = True
        can_edit = False
        hv_profile = get_dormitory_management_profile(request.user)
        if hv_profile:
            from dormitory.models import RoomAssignment as DormRoomAssignment
            from django.db.models import Q
            from datetime import timedelta
            cutoff = today + timedelta(days=90)
            has_access = DormRoomAssignment.objects.filter(
                student=student,
                room__dormitory=hv_profile.dormitory,
                start_date__lte=cutoff,
            ).filter(
                Q(end_date__isnull=True) | Q(end_date__gte=today)
            ).exists()
            if not has_access:
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied
        else:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    elif request.user.is_authenticated and is_travel_expense_office(request.user):
        is_rk_view = True
        can_edit = False

    elif is_ps_role:
        # Ausbildungsverantwortliche: nur freigegebene Nachwuchskräfte
        is_ps_view = True
        can_edit = False
        from student.models import TrainingResponsibleAccess
        if not TrainingResponsibleAccess.objects.filter(user=request.user, student=student).exists():
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    if restricted_view:
        addresses = []
        assignments = []
        custom_fields = []
        paperless_docs = None
    else:
        addresses = [student.address] if student.address else []
        assignments = student.room_assignments.select_related('room__dormitory').order_by('-start_date')
        fv_map = {fv.field_id: fv.value for fv in student.custom_field_values.select_related('field')}
        custom_fields = [
            (defn, fv_map.get(defn.pk, ''))
            for defn in StudentFieldDefinition.objects.all()
        ]
        if is_hv_view or is_rk_view:
            paperless_docs = None
        else:
            paperless_docs = PaperlessService.get_documents_for_student(student.pk)

    grades = list(
        student.grades
        .select_related('grade_type__job_profile')
        .order_by('grade_type__order', 'grade_type__name', '-date')
    ) if not restricted_view and not is_hv_view and not is_rk_view else []

    current_block = None
    remaining_days = None
    if student.course:
        current_block = ScheduleBlock.objects.filter(
            course=student.course,
            start_date__lte=today,
            end_date__gte=today,
        ).first()
        if current_block:
            remaining_days = (current_block.end_date - today).days

    from course.models import InternshipAssignment
    current_internship = InternshipAssignment.objects.filter(
        student=student,
        start_date__lte=today,
        end_date__gte=today,
    ).select_related('unit', 'schedule_block').first()
    internship_remaining_days = (current_internship.end_date - today).days if current_internship else None

    ia_qs = InternshipAssignment.objects.filter(student=student)
    if koord_unit_pks is not None and not is_ps_view:
        ia_qs = ia_qs.filter(unit_id__in=koord_unit_pks)
    internship_assignments = list(
        ia_qs.select_related('unit', 'schedule_block').order_by('-start_date')
    )

    from services.roles import is_training_director
    can_confirm_grades = is_training_director(request.user) or request.user.is_staff

    # Tab-Sichtbarkeits-Flags
    tab_wohnheim = not is_hv_view
    tab_praktika = not is_hv_view
    tab_noten = not restricted_view and not is_hv_view and not is_rk_view
    tab_akte = not is_hv_view and not is_rk_view
    tab_freigaben = False

    # Ausbildungsnachweise
    from services.roles import is_training_director as _is_leitung, is_training_office as _is_referat
    can_view_training = (
        _is_leitung(request.user) or _is_referat(request.user) or is_ps_role
    ) and not restricted_view
    training_records = (
        student.training_records.order_by('-week_start')
        if can_view_training else None
    )

    can_view_contacts = _is_leitung(request.user) or _is_referat(request.user)
    contacts = (
        student.contact_entries.select_related('recorded_by').all()
        if can_view_contacts else None
    )

    can_manage_notes = _is_leitung(request.user) or _is_referat(request.user)
    internal_notes = (
        student.internal_notes.select_related('created_by').all()
        if can_manage_notes else None
    )

    can_manage_checklists = (_is_leitung(request.user) or _is_referat(request.user)) and not restricted_view
    checklists = (
        student.checklists.prefetch_related('items').all()
        if can_manage_checklists else None
    )
    checklist_templates = ChecklistTemplate.objects.filter(is_active=True) if can_manage_checklists else []

    # Lern- und Studientage
    can_view_study_days = (_is_leitung(request.user) or _is_referat(request.user)) and not restricted_view
    study_day_requests = None
    study_day_balance = None
    if can_view_study_days:
        from studyday.models import StudyDayRequest, get_study_day_balance
        study_day_requests = StudyDayRequest.objects.filter(student=student).order_by('-date')
        study_day_balance = get_study_day_balance(student)

    # Abwesenheiten (Urlaub + Krank + Lerntage, kombinierter Tab)
    can_view_absences = (
        _is_leitung(request.user) or _is_referat(request.user) or is_ps_role
    ) and not restricted_view

    # Kalender (nur Leitung / Referat, nicht eingeschränkte Ansicht)
    can_view_calendar = (
        _is_leitung(request.user) or _is_referat(request.user)
    ) and not restricted_view
    calendar_data = None
    if can_view_calendar:
        from student.calendar_utils import build_student_calendar
        cal_year = int(request.GET.get('year', date.today().year))
        calendar_data = build_student_calendar(
            student, cal_year,
            include_interventions=True,
            portal_view=False,
        )

    # Maßnahmen (nur Leitung / Referat, nicht eingeschränkte Ansicht)
    can_view_interventions = (
        _is_leitung(request.user) or _is_referat(request.user)
    ) and not restricted_view
    interventions = (
        student.interventions.select_related('category', 'created_by')
        .order_by('-date')
        if can_view_interventions else None
    )

    # Einzelberechtigungen für Urlaubsgenehmigung / Lerntage
    from services.roles import get_training_office_profile as _get_referat_profile
    _referat_profile = _get_referat_profile(request.user) if _is_referat(request.user) and not _is_leitung(request.user) else None
    can_approve_vacation   = _is_leitung(request.user) or (_referat_profile and _referat_profile.can_approve_vacation)
    can_approve_study_days = _is_leitung(request.user) or (_referat_profile and _referat_profile.can_approve_study_days)
    vacation_requests = None
    sick_leaves = None
    traffic_light_info = None
    if can_view_absences:
        from absence.models import VacationRequest, SickLeave, TRAFFIC_LIGHT_ICON, update_traffic_light
        vacation_requests = (
            VacationRequest.objects
            .filter(student=student)
            .prefetch_related('cancellation_requests')
            .order_by('-start_date')
        )
        sick_leaves = SickLeave.objects.filter(student=student).order_by('-start_date')
        traffic_light = update_traffic_light(student)
        icon_data = TRAFFIC_LIGHT_ICON.get(traffic_light, TRAFFIC_LIGHT_ICON['unknown'])
        traffic_light_info = {
            'value': traffic_light,
            'color':   icon_data[0],
            'icon':    icon_data[1],
            'label':   icon_data[2],
        }

    # Ausbildungsplan (Soll-Ist-Abgleich)
    curriculum_status = []
    if student.course and hasattr(student.course, 'job_profile') and student.course.job_profile:
        from course.curriculum import get_curriculum_status
        curriculum_status = get_curriculum_status(student)

    return render(request, 'student/student_detail.html', {
        'student': student,
        'student_active': _is_active(student),
        'all_statuses': Status.objects.all(),
        'addresses': addresses,
        'assignments': assignments,
        'custom_fields': custom_fields,
        'paperless_docs': paperless_docs,
        'current_block': current_block,
        'remaining_days': remaining_days,
        'current_internship': current_internship,
        'internship_remaining_days': internship_remaining_days,
        'restricted_view': restricted_view,
        'can_edit': can_edit,
        'grades': grades,
        'internship_assignments': internship_assignments,
        'can_confirm_grades': can_confirm_grades,
        'document_templates': (
            StudentDocumentTemplate.objects.filter(is_active=True).prefetch_related('extra_fields')
            if can_edit else []
        ),
        'tab_wohnheim': tab_wohnheim,
        'tab_praktika': tab_praktika,
        'tab_noten': tab_noten,
        'tab_akte': tab_akte,
        'tab_freigaben': tab_freigaben,
        'training_records': training_records,
        'can_review_training': _is_leitung(request.user) or is_ps_role,
        'contacts': contacts,
        'can_view_contacts': can_view_contacts,
        'contact_type_choices': ContactEntry.CONTACT_TYPE_CHOICES,
        'checklists': checklists,
        'checklist_templates': checklist_templates,
        'can_manage_checklists': can_manage_checklists,
        'internal_notes': internal_notes,
        'can_manage_notes': can_manage_notes,
        'study_day_requests': study_day_requests,
        'study_day_balance': study_day_balance,
        'can_view_study_days': can_view_study_days,
        'can_view_absences': can_view_absences,
        'can_approve_vacation': can_approve_vacation,
        'can_approve_study_days': can_approve_study_days,
        'vacation_requests': vacation_requests,
        'sick_leaves': sick_leaves,
        'traffic_light_info': traffic_light_info,
        'interventions': interventions,
        'can_view_interventions': can_view_interventions,
        'calendar_data': calendar_data,
        'can_view_calendar': can_view_calendar,
        'curriculum_status': curriculum_status,
    })


@login_required
@require_POST
def student_document_upload(request, pk):
    from django.core.exceptions import PermissionDenied
    from services.roles import is_training_coordinator
    if is_training_coordinator(request.user):
        raise PermissionDenied
    student = get_object_or_404(Student, pk=pk)
    uploaded_file = request.FILES.get('file')
    title = request.POST.get('title', '').strip()
    if not uploaded_file or not title:
        messages.error(request, 'Bitte Titel und Datei angeben.')
        return redirect(f'/student/{pk}/?tab=akte')
    doc_id = PaperlessService.upload_and_wait(
        file_bytes=uploaded_file.read(),
        title=title,
        student_id=student.pk,
        filename=uploaded_file.name,
        mime_type=uploaded_file.content_type or 'application/octet-stream',
    )
    if doc_id:
        messages.success(request, f'„{title}" wurde erfolgreich in die Akte hochgeladen.')
    else:
        messages.error(request, 'Upload zu Paperless fehlgeschlagen. Bitte erneut versuchen.')
    return redirect(f'/student/{pk}/?tab=akte')


@login_required
def student_document_template_fields(request, template_pk):
    """Felder-Editor (Inline-Formset) für eine Dokumentvorlage."""
    from services.roles import is_training_director
    if not is_training_director(request.user):
        raise PermissionDenied
    template_obj = get_object_or_404(StudentDocumentTemplate, pk=template_pk)
    if request.method == 'POST':
        formset = StudentDocumentTemplateFieldFormSet(request.POST, instance=template_obj)
        if formset.is_valid():
            formset.save()
            messages.success(request, f'Felder für „{template_obj.name}" wurden gespeichert.')
            return redirect('student:student_document_template_fields', template_pk=template_obj.pk)
    else:
        formset = StudentDocumentTemplateFieldFormSet(instance=template_obj)
    return render(request, 'student/template_fields_edit.html', {
        'template_obj': template_obj,
        'formset': formset,
    })


def _parse_extra_field_values(post, extra_fields):
    """Parst die User-Eingaben für Vorlagenfelder. Gibt (values, errors) zurück."""
    from datetime import datetime as _datetime
    values: dict = {}
    errors: dict = {}
    for f in extra_fields:
        raw = (post.get(f'extra_{f.pk}') or '').strip()
        if f.required and not raw:
            errors[f.pk] = 'Pflichtfeld.'
            values[f.key] = ''
            continue
        if not raw:
            values[f.key] = ''
            continue
        if f.field_type == StudentDocumentTemplateField.FIELD_TYPE_DATE:
            try:
                d = _datetime.strptime(raw, '%Y-%m-%d').date()
                values[f.key] = d.strftime('%d.%m.%Y')
            except ValueError:
                errors[f.pk] = 'Ungültiges Datum (erwartet YYYY-MM-DD).'
                values[f.key] = raw
        elif f.field_type == StudentDocumentTemplateField.FIELD_TYPE_NUMBER:
            try:
                num = float(raw.replace(',', '.'))
                values[f.key] = int(num) if num.is_integer() else num
            except ValueError:
                errors[f.pk] = 'Ungültige Zahl.'
                values[f.key] = raw
        elif f.field_type == StudentDocumentTemplateField.FIELD_TYPE_SELECT:
            if raw not in f.get_options_list():
                errors[f.pk] = 'Bitte einen der vorgegebenen Werte wählen.'
            values[f.key] = raw
        else:
            values[f.key] = raw
    return values, errors


@login_required
def student_document_generate(request, pk, template_pk):
    from django.core.exceptions import PermissionDenied
    from services.roles import is_training_coordinator
    if is_training_coordinator(request.user):
        raise PermissionDenied
    from io import BytesIO
    from datetime import date as _date
    student = get_object_or_404(Student, pk=pk)
    template_obj = get_object_or_404(StudentDocumentTemplate, pk=template_pk, is_active=True)
    extra_fields = list(template_obj.extra_fields.all())

    def _render_form(values_post=None, errors=None):
        errors = errors or {}
        items = []
        for f in extra_fields:
            value = ''
            if values_post is not None:
                value = values_post.get(f'extra_{f.pk}', '')
            items.append({'field': f, 'value': value, 'error': errors.get(f.pk)})
        return render(request, 'student/document_generate_form.html', {
            'student': student,
            'template_obj': template_obj,
            'items': items,
        })

    # Vorlage hat Zusatzfelder → GET zeigt Formular
    if extra_fields and request.method == 'GET':
        return _render_form()

    if not extra_fields and request.method != 'POST':
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(['POST'])

    # Zusatzfelder validieren (nur falls vorhanden)
    extra_values: dict = {}
    if extra_fields:
        extra_values, field_errors = _parse_extra_field_values(request.POST, extra_fields)
        if field_errors:
            return _render_form(values_post=request.POST, errors=field_errors)
    from document.contexts import student_context, course_context, creator_context, meta_context
    from document.render import render_docx, upload_to_paperless
    context = {
        **student_context(student),
        **course_context(student.course),
        **creator_context(request.user),
        **meta_context(),
    }
    # Extra-Felder ergänzen (Standard-Platzhalter haben Vorrang)
    for k, v in extra_values.items():
        context.setdefault(k, v)
    try:
        file_bytes = render_docx(template_obj.template_file.path, context)
    except Exception as e:
        messages.error(request, f'Fehler beim Erstellen des Dokuments: {e}')
        return redirect(f'/student/{pk}/?tab=akte')

    today_str = _date.today().strftime('%Y%m%d')
    title = f'{template_obj.name} – {student.first_name} {student.last_name} – {_date.today().strftime("%d.%m.%Y")}'
    filename = f'{template_obj.name}_{student.last_name}_{today_str}.docx'
    doc_id = upload_to_paperless(
        file_bytes=file_bytes,
        title=title,
        student_id=student.pk,
        filename=filename,
    )

    # Audit-Log: Eingaben in die Vorlagenfelder festhalten (sofern welche vorhanden)
    if extra_values:
        try:
            from auditlog.manual import log_event
            from auditlog.models import AuditLogEntry
            log_event(
                action=AuditLogEntry.ACTION_CREATE,
                instance=template_obj,
                user=request.user,
                changes={
                    'student_id':       student.pk,
                    'document_title':   title,
                    'paperless_doc_id': doc_id,
                    'extra_values':     {k: str(v) for k, v in extra_values.items()},
                },
                student_id=student.pk,
            )
        except Exception:
            import logging
            logging.getLogger(__name__).exception('Audit-Log für Dokument-Generierung fehlgeschlagen')

    if doc_id:
        messages.success(request, f'„{template_obj.name}" wurde generiert und in die Akte geladen.')
    else:
        messages.error(request, 'Upload zu Paperless fehlgeschlagen.')
    return redirect(f'/student/{pk}/?tab=akte')



def _get_custom_fields_with_values(student=None):
    """Returns list of (definition, current_value) for the form."""
    definitions = StudentFieldDefinition.objects.all()
    if student is None:
        return [(defn, '') for defn in definitions]
    fv_map = {fv.field_id: fv.value for fv in student.custom_field_values.all()}
    return [(defn, fv_map.get(defn.pk, '')) for defn in definitions]


def _save_custom_fields(post_data, student):
    for defn in StudentFieldDefinition.objects.all():
        value = post_data.get(f'custom_{defn.pk}', '').strip()
        StudentFieldValue.objects.update_or_create(
            student=student, field=defn,
            defaults={'value': value},
        )


def _maybe_create_portal_account(request, student):
    """
    Creates a Django User portal account for a student if:
    - the student has an email_id set
    - no account is linked yet
    """
    import logging
    logger = logging.getLogger(__name__)

    if student.user_id:
        return  # already linked
    if not student.email_id:
        messages.warning(request, f'Kein Portal-Konto angelegt: {student.first_name} {student.last_name} hat keine dienstliche E-Mail-Adresse.')
        return

    from django.contrib.auth.models import User

    email = student.email_id
    username_base = email.split('@')[0]
    username = username_base
    suffix = 1
    while User.objects.filter(username=username).exists():
        username = f'{username_base}{suffix}'
        suffix += 1

    user = User.objects.create_user(
        username=username,
        email=email,
        first_name=student.first_name,
        last_name=student.last_name,
    )
    user.set_unusable_password()
    user.save(update_fields=['password'])
    student.user = user
    student.save(update_fields=['user'])

    # Build one-time password-set link (same pattern as ChiefInstructor welcome)
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.encoding import force_bytes
    from django.utils.http import urlsafe_base64_encode
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    reset_url = request.build_absolute_uri(f'/accounts/reset/{uid}/{token}/')

    try:
        from services.email import send_mail
        from services.models import NotificationTemplate
        portal_url = request.build_absolute_uri('/portal/')
        subject, body = NotificationTemplate.render('student_portal_welcome', {
            'vorname':      student.first_name,
            'nachname':     student.last_name,
            'benutzername': username,
            'reset_url':    reset_url,
            'passwort_url': reset_url,   # Alias für ältere DB-Vorlagen
            'portal_url':   portal_url,  # Alias für ältere DB-Vorlagen
            'passwort':     reset_url,   # weiterer Alias
        })
        send_mail(subject=subject, body_text=body, recipient_list=[email])
        messages.info(request, f'Portal-Konto „{username}" für {student.first_name} {student.last_name} angelegt und Einladungs-E-Mail gesendet.')
    except Exception as exc:
        logger.warning('Welcome-Mail für %s fehlgeschlagen: %s', student.pk, exc)
        messages.warning(request, f'Portal-Konto „{username}" angelegt, aber Einladungs-E-Mail konnte nicht gesendet werden: {exc}')


@login_required
def student_create(request):
    custom_fields = _get_custom_fields_with_values()
    form = StudentForm(request.POST or None)
    address_form = AdressForm(request.POST or None)
    if form.is_valid():
        student = form.save()
        if not student.status_id:
            from student.models import Status
            aktiv = Status.objects.filter(description__iexact='aktiv').first()
            if aktiv:
                student.status = aktiv
                student.save(update_fields=['status'])
        _save_custom_fields(request.POST, student)
        if any(request.POST.get(f) for f in ['street', 'house_number', 'zip_code', 'city']):
            if address_form.is_valid():
                addr = address_form.save()
                student.address = addr
                student.save(update_fields=['address'])
        _maybe_create_portal_account(request, student)
        messages.success(request, f'{student.first_name} {student.last_name} wurde erfolgreich angelegt.')
        return redirect('student:student_detail', pk=student.pk)
    return render(request, 'student/student_form.html', {
        'form': form,
        'address_form': address_form,
        'custom_fields': custom_fields,
        'action': 'Anlegen',
    })


@login_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    custom_fields = _get_custom_fields_with_values(student)
    existing_address = student.address
    form = StudentForm(request.POST or None, instance=student)
    address_form = AdressForm(request.POST or None, instance=existing_address)
    if form.is_valid():
        form.save()
        _save_custom_fields(request.POST, student)
        if address_form.is_valid():
            if any(request.POST.get(f) for f in ['street', 'house_number', 'zip_code', 'city']):
                addr = address_form.save()
                if student.address_id != addr.pk:
                    student.address = addr
                    student.save(update_fields=['address'])
            elif existing_address:
                student.address = None
                student.save(update_fields=['address'])
                existing_address.delete()
        student.refresh_from_db()
        _maybe_create_portal_account(request, student)
        messages.success(request, f'{student.first_name} {student.last_name} wurde erfolgreich gespeichert.')
        return redirect('student:student_detail', pk=student.pk)
    return render(request, 'student/student_form.html', {
        'form': form,
        'address_form': address_form,
        'custom_fields': custom_fields,
        'action': 'Bearbeiten',
        'student': student,
    })

@login_required
def student_statistics(request):
    today = date.today()
    all_students = Student.objects.select_related('gender', 'employment', 'course', 'status').all()
    active = [s for s in all_students if _is_active(s)]
    total_active = len(active)

    # 1.1 Nach Geschlecht
    gender_counts = Counter(
        (s.gender.description if s.gender else 'Nicht angegeben') for s in active
    )

    # 1.2 Nach Beschäftigungsverhältnis
    employment_counts = Counter(
        (s.employment.description if s.employment else 'Nicht angegeben') for s in active
    )

    # 1.3 Nach Alter (Altersgruppen)
    def calc_age(dob):
        return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

    age_brackets = ['unter 20', '20–22', '23–25', '26–28', 'über 28']
    age_counts = Counter()
    for s in active:
        age = calc_age(s.date_of_birth)
        if age < 20:
            age_counts['unter 20'] += 1
        elif age <= 22:
            age_counts['20–22'] += 1
        elif age <= 25:
            age_counts['23–25'] += 1
        elif age <= 28:
            age_counts['26–28'] += 1
        else:
            age_counts['über 28'] += 1

    # 1.4 Nach Kurs
    course_counts = Counter(
        (s.course.title if s.course else 'Kein Kurs') for s in active
    )

    # 1.5 Nach voraussichtlichem Abschlussdatum (Kurs-Enddatum, gruppiert nach Jahr)
    end_date_counts = Counter()
    for s in active:
        if s.course and s.course.end_date:
            end_date_counts[str(s.course.end_date.year)] += 1
        else:
            end_date_counts['Unbekannt'] += 1

    def chart_data(counter, key_order=None):
        keys = key_order if key_order else sorted(counter.keys())
        return {
            'labels': keys,
            'values': [counter.get(k, 0) for k in keys],
        }

    return render(request, 'student/student_statistics.html', {
        'total_active': total_active,
        'total_all': len(all_students),
        'gender': chart_data(gender_counts),
        'employment': chart_data(employment_counts),
        'age': chart_data(age_counts, key_order=age_brackets),
        'course': chart_data(course_counts),
        'end_date': chart_data(end_date_counts),
    })


@login_required
@require_POST
def student_set_status(request, pk):
    """Setzt den Status einer Nachwuchskraft direkt (z. B. über Schnellschaltflächen)."""
    student = get_object_or_404(Student, pk=pk)
    status_pk = request.POST.get('status_pk')
    if status_pk:
        status = get_object_or_404(Status, pk=status_pk)
        student.status = status
        student.save(update_fields=['status', 'status_changed_at'])
        messages.success(request, f'Status von {student.first_name} {student.last_name} auf „{status.description}" gesetzt.')
    return redirect('student:student_detail', pk=student.pk)


@login_required
def data_privacy_overview(request):
    """Übersicht über Datenschutz-/Anonymisierungsstatus der Nachwuchskräfte."""
    due = get_students_due_for_anonymization()
    approaching = get_students_approaching_anonymization()
    anonymized = Student.objects.filter(anonymized_at__isnull=False).order_by('anonymized_at')
    return render(request, 'student/data_privacy.html', {
        'due': due,
        'approaching': approaching,
        'anonymized': anonymized,
    })


@login_required
def student_competence_matrix(request, pk):
    """Kompetenzmatrix-Visualisierung für eine einzelne Nachwuchskraft (Verwaltungssicht).

    Zugriff:
    - Ausbildungsleitung / -referat / Staff: alle Nachwuchskräfte
    - Ausbildungskoordination: nur eigene zugewiesene Nachwuchskräfte
    """
    from services.roles import (
        is_training_director, is_training_office,
        is_training_coordinator, get_chief_instructor,
    )
    from services.competence_matrix import get_competence_matrix
    from instructor.views import _get_coordination_area
    from course.models import InternshipAssignment

    student = get_object_or_404(
        Student.objects.select_related('course__job_profile'),
        pk=pk,
    )

    user = request.user
    can_view_full = (
        user.is_staff
        or is_training_director(user)
        or is_training_office(user)
    )
    if not can_view_full and is_training_coordinator(user):
        chief = get_chief_instructor(user)
        if not (chief and chief.coordination):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        descendant_pks, _, _ = _get_coordination_area(chief.coordination)
        cutoff = date.today() - __import__('datetime').timedelta(days=14)
        has_access = InternshipAssignment.objects.filter(
            student=student, unit_id__in=descendant_pks, end_date__gte=cutoff,
        ).exists()
        if not has_access:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
    elif not can_view_full:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    matrix = get_competence_matrix(student)

    return render(request, 'student/competence_matrix.html', {
        'student': student,
        'matrix':  matrix,
    })


@login_required
@require_POST
def anonymize_student_view(request, pk):
    """Anonymisiert eine einzelne Nachwuchskraft."""
    student = get_object_or_404(Student, pk=pk)
    try:
        anonymize_student(student)
        messages.success(request, f'Nachwuchskraft {pk} wurde erfolgreich anonymisiert.')
    except Exception as exc:
        messages.error(request, f'Fehler bei der Anonymisierung von {pk}: {exc}')
    return redirect('student:data_privacy')


@login_required
@require_POST
def anonymize_all(request):
    """Anonymisiert alle fälligen Nachwuchskräfte auf einmal."""
    students = get_students_due_for_anonymization()
    success_count = 0
    error_count = 0
    for student in students:
        try:
            anonymize_student(student)
            success_count += 1
        except Exception as exc:
            messages.error(request, f'Fehler bei {student.pk}: {exc}')
            error_count += 1
    if success_count:
        messages.success(request, f'{success_count} Nachwuchskraft/-kräfte erfolgreich anonymisiert.')
    if error_count:
        messages.warning(request, f'{error_count} Anonymisierung(en) fehlgeschlagen.')
    return redirect('student:data_privacy')


# ── Noten ──────────────────────────────────────────────────────────────────────

@login_required
def grade_create(request, student_pk):
    student = get_object_or_404(Student.objects.select_related('course__job_profile'), pk=student_pk)
    form = GradeForm(request.POST or None, request.FILES or None, student=student)
    if form.is_valid():
        grade = form.save(commit=False)
        grade.student = student
        grade.save()
        attachment = request.FILES.get('attachment')
        if attachment:
            title = f"{student} – {grade.grade_type.name}"
            doc_id = PaperlessService.upload_and_wait(
                file_bytes=attachment.read(),
                title=title,
                student_id=student.pk,
                filename=attachment.name,
                mime_type=attachment.content_type,
            )
            if doc_id:
                grade.paperless_document_id = doc_id
                grade.save(update_fields=['paperless_document_id'])
                messages.success(request, f'Note gespeichert und Anhang in Paperless abgelegt.')
            else:
                messages.warning(request, f'Note gespeichert, aber Anhang konnte nicht hochgeladen werden.')
        else:
            messages.success(request, f'Note erfolgreich gespeichert.')
        return redirect('student:student_detail', pk=student.pk)
    return render(request, 'student/grade_form.html', {
        'form': form,
        'student': student,
        'action': 'Anlegen',
    })


@login_required
def grade_edit(request, student_pk, grade_public_id):
    student = get_object_or_404(Student.objects.select_related('course__job_profile'), pk=student_pk)
    grade = get_object_or_404(Grade, pk=grade_public_id, student=student)
    form = GradeForm(request.POST or None, request.FILES or None, instance=grade, student=student)
    if form.is_valid():
        grade = form.save()
        attachment = request.FILES.get('attachment')
        if attachment:
            if grade.paperless_document_id:
                PaperlessService.delete_document(grade.paperless_document_id)
            title = f"{student} – {grade.grade_type.name}"
            doc_id = PaperlessService.upload_and_wait(
                file_bytes=attachment.read(),
                title=title,
                student_id=student.pk,
                filename=attachment.name,
                mime_type=attachment.content_type,
            )
            if doc_id:
                grade.paperless_document_id = doc_id
                grade.save(update_fields=['paperless_document_id'])
                messages.success(request, 'Note gespeichert und Anhang aktualisiert.')
            else:
                messages.warning(request, 'Note gespeichert, aber neuer Anhang konnte nicht hochgeladen werden.')
        else:
            messages.success(request, 'Note erfolgreich gespeichert.')
        return redirect('student:student_detail', pk=student.pk)
    return render(request, 'student/grade_form.html', {
        'form': form,
        'student': student,
        'grade': grade,
        'action': 'Bearbeiten',
    })


@login_required
@require_POST
def grade_delete(request, student_pk, grade_public_id):
    student = get_object_or_404(Student, pk=student_pk)
    grade = get_object_or_404(Grade, pk=grade_public_id, student=student)
    if grade.paperless_document_id:
        PaperlessService.delete_document(grade.paperless_document_id)
    grade.delete()
    messages.success(request, 'Note wurde gelöscht.')
    return redirect('student:student_detail', pk=student.pk)


# ── CSV-Export ─────────────────────────────────────────────────────────────────

EXPORT_FIELD_GROUPS = [
    ('Stammdaten', [
        ('id',          'Azubi-ID'),
        ('last_name',   'Nachname'),
        ('first_name',  'Vorname'),
        ('gender',      'Geschlecht'),
        ('date_of_birth',  'Geburtsdatum'),
        ('place_of_birth', 'Geburtsort'),
        ('employment',  'Beschäftigungsverhältnis'),
        ('status',      'Status'),
        ('course',      'Kurs'),
    ]),
    ('Kontakt', [
        ('phone_number',   'Telefonnummer'),
        ('email_private',  'E-Mail (privat)'),
        ('email_id',       'E-Mail (dienstlich)'),
    ]),
    ('Adresse', [
        ('address_street',       'Straße'),
        ('address_house_number', 'Hausnummer'),
        ('address_zip',          'PLZ'),
        ('address_city',         'Ort'),
    ]),
    ('Meta', [
        ('created_at', 'Erstellt am'),
    ]),
]

_FLAT_FIELD_MAP = {key: label for _, fields in EXPORT_FIELD_GROUPS for key, label in fields}


def _get_student_field_value(student, field_key):
    """
    Gibt den formatierten Wert eines Nachwuchskraft-Feldes als Zeichenkette zurück.

    Wird beim CSV-Export verwendet, um Felder flexibel über einen Schlüssel abzurufen.
    Datumswerte werden im Format TT.MM.JJJJ ausgegeben; fehlende Felder ergeben einen
    leeren String, sodass der Export immer eine vollständige Spaltenstruktur liefert.
    """
    if field_key == 'id':
        return student.id
    if field_key == 'last_name':
        return student.last_name
    if field_key == 'first_name':
        return student.first_name
    if field_key == 'gender':
        return str(student.gender) if student.gender else ''
    if field_key == 'date_of_birth':
        return student.date_of_birth.strftime('%d.%m.%Y') if student.date_of_birth else ''
    if field_key == 'place_of_birth':
        return student.place_of_birth
    if field_key == 'employment':
        return str(student.employment) if student.employment else ''
    if field_key == 'status':
        return str(student.status) if student.status else ''
    if field_key == 'course':
        return str(student.course) if student.course else ''
    if field_key == 'phone_number':
        return student.phone_number or ''
    if field_key == 'email_private':
        return student.email_private or ''
    if field_key == 'email_id':
        return student.email_id or ''
    if field_key == 'address_street':
        return student.address.street if student.address else ''
    if field_key == 'address_house_number':
        return student.address.house_number if student.address else ''
    if field_key == 'address_zip':
        return student.address.zip_code if student.address else ''
    if field_key == 'address_city':
        return student.address.city if student.address else ''
    if field_key == 'created_at':
        return student.created_at.strftime('%d.%m.%Y') if student.created_at else ''
    return ''


@login_required
def student_export_csv(request):
    all_statuses = Status.objects.all()
    custom_field_defs = StudentFieldDefinition.objects.all().order_by('name')

    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        selected_fields = request.POST.getlist('fields')

        students = (
            Student.objects
            .filter(pk__in=student_ids)
            .select_related('gender', 'course', 'status', 'employment', 'address')
            .prefetch_related('custom_field_values__field')
            .order_by('last_name', 'first_name')
        )

        custom_field_id_map = {str(f.pk): f.name for f in custom_field_defs}

        headers = []
        for f in selected_fields:
            if f in _FLAT_FIELD_MAP:
                headers.append(_FLAT_FIELD_MAP[f])
            elif f.startswith('custom_'):
                cid = f[len('custom_'):]
                headers.append(custom_field_id_map.get(cid, f))

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="nachwuchskraefte.csv"'
        response.write('\ufeff')  # BOM für Excel
        writer = csv.writer(response, delimiter=';')
        writer.writerow(headers)

        for s in students:
            fv_map = {str(fv.field_id): fv.value for fv in s.custom_field_values.all()}
            row = []
            for f in selected_fields:
                if f in _FLAT_FIELD_MAP:
                    row.append(_get_student_field_value(s, f))
                elif f.startswith('custom_'):
                    cid = f[len('custom_'):]
                    row.append(fv_map.get(cid, ''))
            writer.writerow(row)

        return response

    # GET: Auswahlseite anzeigen
    from course.models import Course as CourseModel
    all_courses = CourseModel.objects.order_by('title')

    status_filter = request.GET.get('status', 'aktiv')
    course_filter = request.GET.get('course', '')

    qs = Student.objects.select_related('gender', 'course', 'status').order_by('last_name', 'first_name')

    if status_filter == 'alle':
        pass
    elif status_filter == 'kein':
        qs = qs.filter(status__isnull=True)
    else:
        qs = qs.filter(status__description__iexact=status_filter)

    if course_filter == 'kein':
        qs = qs.filter(course__isnull=True)
    elif course_filter:
        qs = qs.filter(course__title=course_filter)

    return render(request, 'student/student_export.html', {
        'students': qs,
        'all_statuses': all_statuses,
        'all_courses': all_courses,
        'status_filter': status_filter,
        'course_filter': course_filter,
        'export_field_groups': EXPORT_FIELD_GROUPS,
        'custom_field_defs': custom_field_defs,
    })


@login_required
@require_POST
def contact_entry_create(request, pk):
    from django.core.exceptions import PermissionDenied
    from django.urls import reverse
    from django.utils.dateparse import parse_datetime
    from django.utils.timezone import make_aware
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    student = get_object_or_404(Student, pk=pk)
    detail_url = reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=kontakte'
    contact_type = request.POST.get('contact_type', '')
    inquiry = request.POST.get('inquiry', '').strip()
    response_text = request.POST.get('response', '').strip()
    contacted_at = request.POST.get('contacted_at', '').strip()
    if not inquiry or not contacted_at or contact_type not in dict(ContactEntry.CONTACT_TYPE_CHOICES):
        messages.error(request, 'Bitte alle Pflichtfelder ausfüllen.')
        return redirect(detail_url)
    dt = parse_datetime(contacted_at)
    if dt is None:
        messages.error(request, 'Ungültiges Datum.')
        return redirect(detail_url)
    if dt.tzinfo is None:
        dt = make_aware(dt)
    ContactEntry.objects.create(
        student=student,
        contact_type=contact_type,
        inquiry=inquiry,
        response=response_text,
        contacted_at=dt,
        recorded_by=request.user,
    )
    messages.success(request, 'Kontakteintrag gespeichert.')
    return redirect(detail_url)


@login_required
@require_POST
def contact_entry_delete(request, pk, entry_public_id):
    from django.core.exceptions import PermissionDenied
    from django.urls import reverse
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    entry = get_object_or_404(ContactEntry, pk=entry_public_id, student__pk=pk)
    entry.delete()
    messages.success(request, 'Kontakteintrag gelöscht.')
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=kontakte')


# ── Interne Notizen ─────────────────────────────────────────────────────────

@login_required
@require_POST
def internal_note_create(request, pk):
    from django.core.exceptions import PermissionDenied
    from django.urls import reverse
    from services.roles import is_training_director, is_training_office
    from .models import InternalNote
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    student = get_object_or_404(Student, pk=pk)
    text = request.POST.get('text', '').strip()
    is_pinned = request.POST.get('is_pinned') == '1'
    if not text:
        messages.error(request, 'Die Notiz darf nicht leer sein.')
        return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=notizen')
    InternalNote.objects.create(
        student=student,
        text=text,
        is_pinned=is_pinned,
        created_by=request.user,
    )
    messages.success(request, 'Notiz gespeichert.')
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=notizen')


@login_required
@require_POST
def internal_note_delete(request, pk, note_public_id):
    from django.core.exceptions import PermissionDenied
    from django.urls import reverse
    from services.roles import is_training_director, is_training_office
    from .models import InternalNote
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    note = get_object_or_404(InternalNote, pk=note_public_id, student__pk=pk)
    note.delete()
    messages.success(request, 'Notiz gelöscht.')
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=notizen')


@login_required
@require_POST
def internal_note_toggle_pin(request, pk, note_public_id):
    from django.core.exceptions import PermissionDenied
    from django.urls import reverse
    from services.roles import is_training_director, is_training_office
    from .models import InternalNote
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    note = get_object_or_404(InternalNote, pk=note_public_id, student__pk=pk)
    note.is_pinned = not note.is_pinned
    note.save(update_fields=['is_pinned'])
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=notizen')


# ── Import ──────────────────────────────────────────────────────────────────


def _parse_import_file(f):
    """CSV oder xlsx parsen → (rows, errors). rows sind JSON-serialisierbare Dicts."""
    name = f.name.lower()
    if name.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
    else:
        try:
            content = f.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            f.seek(0)
            content = f.read().decode('latin-1')
        reader = csv.reader(io.StringIO(content), delimiter=';')
        raw_rows = list(reader)

    if not raw_rows:
        return [], [{'row': 0, 'messages': ['Die Datei ist leer.']}]

    header = [str(h).strip().lower() for h in raw_rows[0]]
    data_rows = raw_rows[1:]

    from services.models import Gender
    from course.models import Course as CourseModel
    from .models import Employment, Status as StudentStatus

    genders = {g.abbreviation.strip().lower(): g for g in Gender.objects.all()}
    courses = {c.title.strip().lower(): c for c in CourseModel.objects.all()}
    employments = {e.description.strip().lower(): e for e in Employment.objects.all()}
    statuses = {s.description.strip().lower(): s for s in StudentStatus.objects.all()}

    rows = []
    errors = []

    for i, raw in enumerate(data_rows, start=2):
        if not any(v for v in raw if v):
            continue  # Leerzeile überspringen
        row = {
            header[j]: str(v).strip() if v is not None else ''
            for j, v in enumerate(raw)
            if j < len(header)
        }
        row_errors = []

        dob_raw = row.get('geburtsdatum', '')
        dob = None
        for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'):
            try:
                dob = datetime.strptime(dob_raw, fmt).date()
                break
            except ValueError:
                pass
        if dob is None:
            row_errors.append(f'Ungültiges Geburtsdatum: {dob_raw!r} (erwartet TT.MM.JJJJ)')

        gender_obj = genders.get(row.get('geschlecht', '').strip().lower())
        if not gender_obj:
            row_errors.append(f'Unbekanntes Geschlecht: {row.get("geschlecht")!r} '
                              f'(verfügbar: {", ".join(genders.keys())})')

        course_obj = courses.get(row.get('kurs', '').strip().lower())
        if not course_obj:
            row_errors.append(f'Unbekannter Kurs: {row.get("kurs")!r}')

        employment_obj = employments.get(row.get('beschaeftigungsverhaeltnis', '').strip().lower())
        if not employment_obj:
            row_errors.append(
                f'Unbekanntes Beschäftigungsverhältnis: {row.get("beschaeftigungsverhaeltnis")!r}'
            )

        if not row.get('vorname', '').strip():
            row_errors.append('Vorname fehlt.')
        if not row.get('nachname', '').strip():
            row_errors.append('Nachname fehlt.')

        status_obj = statuses.get(row.get('status', '').strip().lower())

        # Adresse: alle vier Felder müssen entweder leer oder gefüllt sein
        addr_street = row.get('strasse', '').strip()
        addr_house = row.get('hausnummer', '').strip()
        addr_zip = row.get('plz', '').strip()
        addr_city = row.get('ort', '').strip()
        addr_filled = [bool(x) for x in (addr_street, addr_house, addr_zip, addr_city)]
        has_address = False
        if any(addr_filled):
            if not all(addr_filled):
                row_errors.append(
                    'Adresse unvollständig – bitte Strasse, Hausnummer, PLZ und Ort '
                    'gemeinsam ausfüllen oder alle vier Felder leer lassen.'
                )
            else:
                has_address = True

        parsed = {
            'row_num': i,
            'first_name': row.get('vorname', ''),
            'last_name': row.get('nachname', ''),
            'date_of_birth': dob.isoformat() if dob else None,
            'place_of_birth': row.get('geburtsort', ''),
            'gender_id': gender_obj.pk if gender_obj else None,
            'email_private': row.get('email_privat', '') or None,
            'email_id': row.get('email_kennung', '') or None,
            'phone_number': row.get('telefon', '').replace(' ', '') or None,
            'course_id': course_obj.pk if course_obj else None,
            'employment_id': employment_obj.pk if employment_obj else None,
            'status_id': status_obj.pk if status_obj else None,
            'address_street': addr_street,
            'address_house_number': addr_house,
            'address_zip': addr_zip,
            'address_city': addr_city,
            'has_address': has_address,
            'errors': row_errors,
        }
        rows.append(parsed)
        if row_errors:
            errors.append({'row': i, 'messages': row_errors})

    return rows, errors


def _commit_import(rows):
    from services.models import Adress
    from .models import create_student_id
    count = 0
    for r in rows:
        if r.get('errors'):
            continue
        if not r['first_name'] or not r['last_name'] or not r['date_of_birth']:
            continue
        dob = date.fromisoformat(r['date_of_birth'])
        address = None
        if r.get('has_address'):
            address = Adress.objects.create(
                street=r['address_street'],
                house_number=r['address_house_number'],
                zip_code=r['address_zip'],
                city=r['address_city'],
            )
        Student.objects.create(
            id=create_student_id(),
            first_name=r['first_name'],
            last_name=r['last_name'],
            date_of_birth=dob,
            place_of_birth=r['place_of_birth'],
            gender_id=r['gender_id'],
            email_private=r['email_private'],
            email_id=r['email_id'],
            phone_number=r['phone_number'],
            course_id=r['course_id'],
            employment_id=r['employment_id'],
            status_id=r['status_id'],
            address=address,
        )
        count += 1
    return count


@login_required
def student_import_template(request):
    """Erzeugt eine Excel-Vorlage mit Spalten, Beispielzeile und Referenzlisten."""
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from services.models import Gender
    from course.models import Course as CourseModel
    from .models import Employment, Status as StudentStatus

    wb = openpyxl.Workbook()

    # ── Sheet „Import" ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Import"
    headers = [
        "Vorname", "Nachname", "Geburtsdatum", "Geburtsort", "Geschlecht",
        "Kurs", "Beschaeftigungsverhaeltnis",
        "Email_privat", "Email_Kennung", "Telefon", "Status",
        "Strasse", "Hausnummer", "PLZ", "Ort",
    ]
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [16, 18, 14, 18, 12, 28, 28, 25, 22, 16, 16, 22, 12, 8, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    example = [
        "Max", "Mustermann", "01.07.2005", "Musterstadt", "m",
        "<Kurs-Titel aus Referenzliste>", "<Beschäftigungsverhältnis>",
        "max@privat.de", "MUSTM", "017012345678", "aktiv",
        "Musterstraße", "12a", "12345", "Musterstadt",
    ]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(italic=True, color="888888")

    # Drop-down Geschlecht (Spalte E)
    genders = list(Gender.objects.all())
    if genders:
        gender_values = ",".join(g.abbreviation for g in genders)
        if len(gender_values) <= 250:
            gdv = DataValidation(type="list", formula1=f'"{gender_values}"', allow_blank=False)
            ws.add_data_validation(gdv)
            gdv.add("E2:E1000")

    # Drop-down Status (Spalte K)
    statuses = list(StudentStatus.objects.order_by("description"))
    if statuses:
        status_values = ",".join(s.description for s in statuses)
        if len(status_values) <= 250:
            sdv = DataValidation(type="list", formula1=f'"{status_values}"', allow_blank=True)
            ws.add_data_validation(sdv)
            sdv.add("K2:K1000")

    # ── Referenz-Sheets ─────────────────────────────────────────────────
    def _ref_sheet(title, header_titles, rows_iter, col_widths):
        sh = wb.create_sheet(title)
        for col_idx, h in enumerate(header_titles, start=1):
            c = sh.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
        for col_letter, w in zip("ABCDEFG", col_widths):
            sh.column_dimensions[col_letter].width = w
        sh.freeze_panes = "A2"
        for i, row in enumerate(rows_iter, start=2):
            for col_idx, value in enumerate(row, start=1):
                sh.cell(row=i, column=col_idx, value=value)

    _ref_sheet(
        "Kurse (Referenz)",
        ["Titel"],
        ((c.title,) for c in CourseModel.objects.order_by("title")),
        [60],
    )
    _ref_sheet(
        "Beschäftigungsverh. (Referenz)",
        ["Beschreibung"],
        ((e.description,) for e in Employment.objects.order_by("description")),
        [60],
    )
    _ref_sheet(
        "Status (Referenz)",
        ["Beschreibung"],
        ((s.description,) for s in statuses),
        [40],
    )
    _ref_sheet(
        "Geschlechter (Referenz)",
        ["Abkürzung", "Geschlecht"],
        ((g.abbreviation, g.gender) for g in genders),
        [14, 30],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="nachwuchskraefte_vorlage.xlsx"'
    return response


@login_required
def student_import(request):
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied

    if request.method == 'POST' and 'confirm' in request.POST:
        rows = request.session.pop('import_rows', [])
        if not rows:
            messages.error(request, 'Keine Vorschau-Daten gefunden. Bitte erneut hochladen.')
            return redirect('student:student_import')
        created = _commit_import(rows)
        messages.success(request, f'{created} Nachwuchskraft/kräfte erfolgreich importiert.')
        return redirect('student:student_list')

    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            rows, errors = _parse_import_file(request.FILES['file'])
            if not rows:
                messages.error(request, 'Die Datei enthält keine Daten.')
                return render(request, 'student/student_import.html', {'form': form})
            request.session['import_rows'] = rows
            valid_count = sum(1 for r in rows if not r['errors'])
            return render(request, 'student/student_import_preview.html', {
                'rows': rows,
                'errors': errors,
                'valid_count': valid_count,
            })
    else:
        form = StudentImportForm()

    return render(request, 'student/student_import.html', {'form': form})


# ── Checklisten ────────────────────────────────────────────────────────────────

@login_required
@require_POST
def checklist_create(request, pk):
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    student = get_object_or_404(Student, pk=pk)
    template_pk = request.POST.get('template')
    if template_pk:
        template = get_object_or_404(ChecklistTemplate, pk=template_pk, is_active=True)
        checklist = StudentChecklist.objects.create(
            student=student,
            template=template,
            name=template.name,
            created_by=request.user,
        )
        for item in template.items.order_by('order', 'text'):
            StudentChecklistItem.objects.create(
                checklist=checklist,
                text=item.text,
                order=item.order,
            )
        messages.success(request, f'Checkliste „{checklist.name}" erstellt.')
    else:
        messages.error(request, 'Bitte eine Vorlage auswählen.')
    from django.urls import reverse
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=checklisten')


@login_required
@require_POST
def checklist_item_toggle(request, pk, checklist_public_id, item_public_id):
    from django.utils import timezone
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    student = get_object_or_404(Student, pk=pk)
    item = get_object_or_404(StudentChecklistItem, pk=item_public_id, checklist_id=checklist_public_id, checklist__student=student)
    item.completed = not item.completed
    if item.completed:
        item.completed_at = timezone.now()
        item.completed_by = request.user
    else:
        item.completed_at = None
        item.completed_by = None
    item.save()
    from django.urls import reverse
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=checklisten')


@login_required
@require_POST
def checklist_delete(request, pk, checklist_public_id):
    from services.roles import is_training_director, is_training_office
    if not (is_training_director(request.user) or is_training_office(request.user)):
        raise PermissionDenied
    student = get_object_or_404(Student, pk=pk)
    checklist = get_object_or_404(StudentChecklist, pk=checklist_public_id, student=student)
    checklist.delete()
    messages.success(request, 'Checkliste gelöscht.')
    from django.urls import reverse
    return redirect(reverse('student:student_detail', kwargs={'pk': pk}) + '?tab=checklisten')


@login_required
def coordinator_calendar(request):
    """Übersichtskalender aller Azubis eines Kurses (Koordinationsansicht)."""
    if hasattr(request.user, 'student_profile'):
        raise PermissionDenied

    from course.models import Course
    from student.calendar_utils import build_course_calendar

    courses = Course.objects.order_by('-start_date')
    year = int(request.GET.get('year', date.today().year))
    course_id = request.GET.get('course', '')

    selected_course = None
    course_cal = None

    if course_id:
        selected_course = get_object_or_404(Course, pk=course_id)
        course_cal = build_course_calendar(selected_course, year, include_interventions=True)

    return render(request, 'student/coordinator_calendar.html', {
        'courses': courses,
        'selected_course': selected_course,
        'course_cal': course_cal,
        'year': year,
    })


# ── Ausbildungsplan: Manuelle Bestätigung ────────────────────────────────────

@login_required
@require_POST
def curriculum_toggle_completion(request, pk, requirement_public_id):
    """Ausbildungsplan-Anforderung manuell als erledigt markieren oder Markierung entfernen."""
    from services.roles import is_training_coordinator
    if is_training_coordinator(request.user):
        raise PermissionDenied

    from course.models import CurriculumRequirement, CurriculumCompletion

    student = get_object_or_404(Student, pk=pk)
    requirement = get_object_or_404(CurriculumRequirement, pk=requirement_public_id)

    existing = CurriculumCompletion.objects.filter(student=student, requirement=requirement).first()
    if existing:
        existing.delete()
        messages.success(request, f'Bestätigung für „{requirement.name}" wurde entfernt.')
    else:
        notes = request.POST.get('notes', '').strip()
        CurriculumCompletion.objects.create(
            student=student,
            requirement=requirement,
            completed_by=request.user,
            notes=notes,
        )
        messages.success(request, f'„{requirement.name}" wurde als erledigt markiert.')

    return redirect(f'/student/{pk}/?tab=ausbildungsplan')
